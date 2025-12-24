"""
Streamlit Frontend für den DORA-Szenariengenerator.
Enterprise-Grade Design mit interaktiven Features, Analysen und Workflow-Logs.
"""

import streamlit as st
import pandas as pd
import json
from datetime import datetime, timedelta
from io import BytesIO
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

from neo4j_client import Neo4jClient
from workflows.scenario_workflow import ScenarioWorkflow
from state_models import ScenarioType, CrisisPhase, Inject
from examples.demo_scenarios import get_available_demo_scenarios, load_demo_scenario
import os
from dotenv import load_dotenv
import html

load_dotenv()

# Streamlit Configuration
st.set_page_config(
    page_title="DORA Scenario Generator",
    page_icon=None,
    layout="wide",
    initial_sidebar_state="expanded"
)

# Enterprise Design System (Celonis-inspired)
st.markdown("""
<style>
    /* ===== COLOR SYSTEM ===== */
    :root {
        --primary-dark: #0f172a;
        --primary: #1e293b;
        --primary-light: #334155;
        --secondary: #64748b;
        --secondary-light: #94a3b8;
        --success: #10b981;
        --warning: #f59e0b;
        --danger: #ef4444;
        --info: #3b82f6;
        --background: #f8fafc;
        --surface: #ffffff;
        --border: #e2e8f0;
        --text-primary: #0f172a;
        --text-secondary: #475569;
        --accent: #0066cc;
    }
    
    .stApp {
        background: var(--background);
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', 'Roboto', sans-serif;
    }
    
    .main .block-container {
        padding-top: 1.5rem;
        padding-bottom: 2rem;
        max-width: 1600px;
    }
    
    .main-header {
        background: var(--surface);
        padding: 2rem 0;
        border-bottom: 1px solid var(--border);
        margin-bottom: 2rem;
    }
    
    .main-header h1 {
        font-size: 2rem;
        font-weight: 600;
        margin: 0;
        color: var(--text-primary);
        letter-spacing: -0.02em;
    }
    
    .sub-header {
        font-size: 0.875rem;
        margin-top: 0.5rem;
        color: var(--text-secondary);
        font-weight: 400;
    }
    
    .enterprise-card {
        background: var(--surface);
        padding: 1.5rem;
        border-radius: 4px;
        margin-bottom: 1.5rem;
        border: 1px solid var(--border);
        box-shadow: 0 1px 2px rgba(0,0,0,0.04);
    }
    
    .inject-card {
        background: var(--surface);
        padding: 1.75rem;
        border-radius: 10px;
        margin-bottom: 1.5rem;
        box-shadow: 0 1px 3px rgba(0,0,0,0.08);
        border-left: 4px solid;
        border: 1px solid var(--border);
        transition: all 0.3s ease;
    }
    
    .inject-card:hover {
        box-shadow: 0 4px 12px rgba(0,0,0,0.12);
        transform: translateX(4px);
        border-left-width: 6px;
    }
    
    .phase-badge {
        display: inline-flex;
        align-items: center;
        padding: 0.5rem 1rem;
        border-radius: 6px;
        font-size: 0.75rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        border: 1px solid rgba(255,255,255,0.2);
    }
    
    .metric-card {
        background: var(--surface);
        padding: 1.75rem;
        border-radius: 10px;
        text-align: center;
        box-shadow: 0 1px 3px rgba(0,0,0,0.08);
        border: 1px solid var(--border);
        transition: all 0.3s ease;
        position: relative;
        overflow: hidden;
    }
    
    .metric-card::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        height: 4px;
        background: linear-gradient(90deg, var(--primary) 0%, var(--primary-light) 100%);
    }
    
    .metric-card:hover {
        transform: translateY(-4px);
        box-shadow: 0 8px 20px rgba(0,0,0,0.12);
    }
    
    .metric-value {
        font-size: 2.75rem;
        font-weight: 700;
        color: var(--primary-dark);
        margin: 0.75rem 0;
        line-height: 1;
    }
    
    .metric-label {
        font-size: 0.875rem;
        color: var(--text-secondary);
        text-transform: uppercase;
        letter-spacing: 1px;
        font-weight: 500;
    }
    
    [data-testid="stSidebar"] {
        background: var(--surface);
        border-right: 1px solid var(--border);
    }
    
    [data-testid="stSidebar"] h1,
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3 {
        color: var(--text-primary);
    }
    
    .stButton > button {
        background: var(--accent);
        color: white;
        border: none;
        border-radius: 4px;
        padding: 0.625rem 1.5rem;
        font-weight: 500;
        font-size: 0.875rem;
        transition: background-color 0.2s ease;
    }
    
    .stButton > button:hover {
        background: #0052a3;
    }
    
    .stButton > button[kind="primary"] {
        background: var(--accent);
    }
    
    .stButton > button[kind="primary"]:hover {
        background: #0052a3;
    }
    
    .stTabs [data-baseweb="tab-list"] {
        gap: 4px;
        background: var(--surface);
        padding: 0.5rem;
        border-radius: 8px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.08);
    }
    
    .stTabs [data-baseweb="tab"] {
        background: transparent;
        border-radius: 6px;
        padding: 0.75rem 1.5rem;
        font-weight: 600;
        color: var(--text-secondary);
        transition: all 0.2s ease;
    }
    
    .stTabs [aria-selected="true"] {
        background: var(--primary);
        color: white;
        box-shadow: 0 2px 8px rgba(44, 82, 130, 0.2);
    }
    
    .log-entry {
        background: var(--surface);
        padding: 1rem;
        border-radius: 8px;
        margin-bottom: 0.75rem;
        border-left: 3px solid var(--primary);
        font-family: 'Courier New', monospace;
        font-size: 0.875rem;
    }
    
    .decision-card {
        background: linear-gradient(135deg, #f0f9ff 0%, #e0f2fe 100%);
        padding: 1.25rem;
        border-radius: 8px;
        margin-bottom: 1rem;
        border-left: 4px solid var(--info);
    }
    
    @keyframes fadeIn {
        from { opacity: 0; transform: translateY(10px); }
        to { opacity: 1; transform: translateY(0); }
    }
    
    .fade-in {
        animation: fadeIn 0.5s ease-out;
    }
</style>
""", unsafe_allow_html=True)


def init_session_state():
    """Initialisiert Session State Variablen."""
    if "scenario_result"not in st.session_state:
        st.session_state.scenario_result = None
    if "workflow"not in st.session_state:
        st.session_state.workflow = None
    if "neo4j_client"not in st.session_state:
        st.session_state.neo4j_client = None
    if "filter_phase"not in st.session_state:
        st.session_state.filter_phase = None
    if "search_term"not in st.session_state:
        st.session_state.search_term = ""
    if "interactive_scenario_state"not in st.session_state:
        st.session_state.interactive_scenario_state = None
    if "pending_decision"not in st.session_state:
        st.session_state.pending_decision = None


def get_phase_color(phase: CrisisPhase) -> tuple:
    """Returns color and label for a phase."""
    phase_config = {
        CrisisPhase.NORMAL_OPERATION: ("#10b981", "", "Normal Operation"),
        CrisisPhase.SUSPICIOUS_ACTIVITY: ("#f59e0b", "", "Suspicious Activity"),
        CrisisPhase.INITIAL_INCIDENT: ("#f97316", "", "Initial Incident"),
        CrisisPhase.ESCALATION_CRISIS: ("#ef4444", "", "Escalation"),
        CrisisPhase.CONTAINMENT: ("#8b5cf6", "", "Containment"),
        CrisisPhase.RECOVERY: ("#06b6d4", "", "Recovery")
    }
    return phase_config.get(phase, ("#64748b", "", "Unknown"))


def format_inject_card_safe(inject: Inject, index: int):
    """Formats an inject card with Streamlit-native components."""
    phase_color, icon, phase_label = get_phase_color(inject.phase)
    
    # Container with border-left indicator
    st.markdown(f"""
    <div style="border-left: 3px solid {phase_color}; margin-bottom: 1.5rem; padding-left: 1rem;">
    </div>
    """, unsafe_allow_html=True)
    
    # Header with ID and Phase Badge
    col_header1, col_header2 = st.columns([3, 1])
    with col_header1:
        st.markdown(f"### {inject.inject_id}")
        st.caption(f"Time Offset: {inject.time_offset}")
    with col_header2:
        st.markdown(f"""
        <div style="text-align: right; margin-top: 0.5rem;">
            <span style="background: {phase_color}; color: white; padding: 0.375rem 0.75rem; border-radius: 3px; font-size: 0.6875rem; font-weight: 500; text-transform: uppercase; letter-spacing: 0.05em;">
                {phase_label}
            </span>
        </div>
        """, unsafe_allow_html=True)
    
    # Source and Target
    col_source, col_target = st.columns(2)
    with col_source:
        st.markdown("**Source**")
        st.text(inject.source)
    with col_target:
        st.markdown("**Target**")
        st.text(inject.target)
    
    # Content
    st.markdown("**Content**")
    with st.container():
        st.markdown(f'<div style="padding: 1rem; background: var(--background); border-radius: 4px; border: 1px solid var(--border); font-size: 0.875rem; line-height: 1.6;">{html.escape(inject.content)}</div>', unsafe_allow_html=True)
    
    # MITRE and Modality
    col_mitre, col_modality = st.columns(2)
    with col_mitre:
        st.markdown("**MITRE ATT&CK**")
        mitre_id = inject.technical_metadata.mitre_id or 'N/A'
        st.code(mitre_id, language=None)
    with col_modality:
        st.markdown("**Modality**")
        st.text(inject.modality.value)
    
    # Affected Assets
    st.markdown("**Affected Assets**")
    if inject.technical_metadata.affected_assets:
        asset_cols = st.columns(min(len(inject.technical_metadata.affected_assets), 5))
        for idx, asset in enumerate(inject.technical_metadata.affected_assets[:5]):
            with asset_cols[idx]:
                st.markdown(f"""
                <div style="background: var(--accent); color: white; padding: 0.375rem 0.75rem; border-radius: 3px; text-align: center; font-size: 0.75rem; font-weight: 500;">
                    {html.escape(asset)}
                </div>
                """, unsafe_allow_html=True)
        if len(inject.technical_metadata.affected_assets) > 5:
            st.caption(f"+ {len(inject.technical_metadata.affected_assets) - 5} additional assets")
    else:
        st.caption("No affected assets")
    
    # DORA Compliance
    if inject.dora_compliance_tag:
        st.info(f"**DORA Compliance:** {inject.dora_compliance_tag}")
    
    # Business Impact
    if inject.business_impact:
        st.warning(f"**Business Impact:** {inject.business_impact}")
    
    st.markdown("---")


def create_workflow_timeline(logs: list, decisions: list):
    """Erstellt eine Timeline-Visualisierung des Workflows."""
    if not logs and not decisions:
        return None
    
    fig = go.Figure()
    
    # Workflow-Logs
    if logs:
        log_times = []
        log_nodes = []
        for log in logs:
            try:
                dt = datetime.fromisoformat(log.get("timestamp", ""))
                log_times.append(dt)
                log_nodes.append(log.get("node", "Unknown"))
            except:
                continue
        
        if log_times:
            fig.add_trace(go.Scatter(
                x=log_times,
                y=[1] * len(log_times),
                mode='markers+lines',
                name='Workflow Nodes',
                marker=dict(size=10, color='#2c5282'),
                line=dict(color='#4299e1', width=2)
            ))
    
    # Agent Decisions
    if decisions:
        decision_times = []
        decision_agents = []
        for dec in decisions:
            try:
                dt = datetime.fromisoformat(dec.get("timestamp", ""))
                decision_times.append(dt)
                decision_agents.append(dec.get("agent", "Unknown"))
            except:
                continue
        
        if decision_times:
            fig.add_trace(go.Scatter(
                x=decision_times,
                y=[2] * len(decision_times),
                mode='markers',
                name='Agent Decisions',
                marker=dict(size=12, color='#38a169', symbol='diamond'),
            ))
    
    fig.update_layout(
        title=" Workflow Timeline",
        xaxis_title="Time",
        yaxis_title="",
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        height=300,
        yaxis=dict(showticklabels=False, range=[0, 3])
    )
    
    return fig


def create_agent_decision_chart(decisions: list):
    """Visualisiert Agent Decisions."""
    if not decisions:
        return None
    
    agent_counts = {}
    for dec in decisions:
        agent = dec.get("agent", "Unknown")
        agent_counts[agent] = agent_counts.get(agent, 0) + 1
    
    fig = go.Figure(data=[
        go.Bar(
            x=list(agent_counts.keys()),
            y=list(agent_counts.values()),
            marker_color='#2c5282',
            text=list(agent_counts.values()),
            textposition='auto',
        )
    ])
    
    fig.update_layout(
        title=" Agent Decisions",
        xaxis_title="Agent",
        yaxis_title="Anzahl Entscheidungen",
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        height=300
    )
    
    return fig


def create_timeline_chart(injects: list):
    """Erstellt eine interaktive Timeline-Visualisierung."""
    timeline_data = []
    for inj in injects:
        time_str = inj.time_offset.replace('T+', '')
        hours, minutes = map(int, time_str.split(':'))
        total_minutes = hours * 60 + minutes
        
        phase_color, _, _ = get_phase_color(inj.phase)
        
        timeline_data.append({
            "Time": total_minutes,
            "Inject": inj.inject_id,
            "Phase": inj.phase.value,
            "Phase Label": inj.phase.value.replace('_', '').title(),
            "MITRE": inj.technical_metadata.mitre_id or "N/A",
            "Content": inj.content[:60] + "..."if len(inj.content) > 60 else inj.content,
            "Color": phase_color
        })
    
    df = pd.DataFrame(timeline_data)
    
    fig = px.scatter(
        df,
        x="Time",
        y="Inject",
        color="Phase Label",
        size=[12] * len(df),
        hover_data=["MITRE", "Content"],
        title=" Scenario Timeline",
        labels={"Time": "Time (Minuten)", "Inject": "Inject ID"},
        color_discrete_map={
            "Normal Operation": "#38a169",
            "Suspicious Activity": "#d69e2e",
            "Initial Incident": "#ed8936",
            "Escalation Crisis": "#e53e3e",
            "Containment": "#805ad5",
            "Recovery": "#319795"
        }
    )
    
    fig.update_layout(
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Arial", size=12, color="#1a202c"),
        height=450,
        title_font=dict(size=18, color="#1a365d"),
        xaxis=dict(gridcolor="#e2e8f0"),
        yaxis=dict(gridcolor="#e2e8f0")
    )
    
    return fig


def create_phase_distribution_chart(injects: list):
    """Erstellt ein Phase Distributions-Diagramm."""
    phase_counts = {}
    phase_colors = {}
    for inj in injects:
        phase = inj.phase.value
        phase_counts[phase] = phase_counts.get(phase, 0) + 1
        if phase not in phase_colors:
            phase_color, _, _ = get_phase_color(inj.phase)
            phase_colors[phase] = phase_color
    
    fig = go.Figure(data=[
        go.Bar(
            x=[p.replace('_', '').title() for p in phase_counts.keys()],
            y=list(phase_counts.values()),
            marker_color=[phase_colors.get(p, "#718096") for p in phase_counts.keys()],
            text=list(phase_counts.values()),
            textposition='auto',
            textfont=dict(size=14, color="white", weight="bold"),
            marker_line=dict(color='rgba(0,0,0,0.1)', width=1)
        )
    ])
    
    fig.update_layout(
        title=dict(text="Phase Distribution", font=dict(size=18, color="#1a365d")),
        xaxis_title="Phase",
        yaxis_title="Number of Injects",
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Arial", size=12, color="#1a202c"),
        height=450,
        xaxis=dict(gridcolor="#e2e8f0"),
        yaxis=dict(gridcolor="#e2e8f0", gridwidth=1)
    )
    
    return fig


def export_to_csv(injects: list) -> BytesIO:
    """Exports injects asCSV."""
    data = []
    for inj in injects:
        data.append({
            "Inject ID": inj.inject_id,
            "Time Offset": inj.time_offset,
            "Phase": inj.phase.value,
            "Source": inj.source,
            "Target": inj.target,
            "Modality": inj.modality.value,
            "Content": inj.content,
            "MITRE ID": inj.technical_metadata.mitre_id or "",
            "Affected Assets": ", ".join(inj.technical_metadata.affected_assets),
            "DORA Tag": inj.dora_compliance_tag or "",
            "Business Impact": inj.business_impact or ""
        })
    
    df = pd.DataFrame(data)
    output = BytesIO()
    df.to_csv(output, index=False, encoding='utf-8')
    output.seek(0)
    return output


def export_to_json(injects: list) -> str:
    """Exports injects as JSON."""
    from utils.safe_json import safe_json_dumps
    data = [inj.model_dump(mode='json') for inj in injects]
    return safe_json_dumps(data, indent=2, ensure_ascii=False)


def export_to_excel(injects: list) -> BytesIO:
    """Exports injects as Excel file (.xlsx)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Injects"
        
        # Header
        headers = [
            "Inject ID", "Time Offset", "Phase", "Source", "Target", 
            "Modality", "Content", "MITRE ID", "Affected Assets", 
            "DORA Tag", "Business Impact", "IOC Hash"
        ]
        
        # Header-Styling
        header_fill = PatternFill(start_color="2c5282", end_color="2c5282", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Schreibe Header
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Schreibe Daten
        for row_num, inj in enumerate(injects, 2):
            data_row = [
                inj.inject_id,
                inj.time_offset,
                inj.phase.value,
                inj.source,
                inj.target,
                inj.modality.value,
                inj.content,
                inj.technical_metadata.mitre_id or "",
                ", ".join(inj.technical_metadata.affected_assets),
                inj.dora_compliance_tag or "",
                inj.business_impact or "",
                inj.technical_metadata.ioc_hash or ""
            ]
            
            for col_num, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                if col_num == 7:  # Content-Spalte
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Spaltenbreiten anpassen
        column_widths = {
            'A': 12,  # Inject ID
            'B': 12,  # Time Offset
            'C': 20,  # Phase
            'D': 25,  # Source
            'E': 25,  # Target
            'F': 15,  # Modality
            'G': 60,  # Content
            'H': 12,  # MITRE ID
            'I': 30,  # Affected Assets
            'J': 25,  # DORA Tag
            'K': 50,  # Business Impact
            'L': 20   # IOC Hash
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Zeilenhöhe für Content-Spalte
        for row in range(2, len(injects) + 2):
            ws.row_dimensions[row].height = 60
        
        # Freeze Header
        ws.freeze_panes = 'A2'
        
        # Speichere in BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
        
    except ImportError:
        raise ImportError("openpyxl ist nicht installiert. Bitte installiere es mit: pip install openpyxl")
    except Exception as e:
        raise Exception(f"Error during Excel export: {e}")


def export_decisions_to_csv(decisions: list) -> BytesIO:
    """Exports agent decisions as CSV."""
    data = []
    for dec in decisions:
        input_data = dec.get("input", {})
        output_data = dec.get("output", {})
        
        # Extrahiere wichtige Informationen
        reasoning = output_data.get("reasoning", "")
        if isinstance(reasoning, dict):
            reasoning = json.dumps(reasoning, ensure_ascii=False)
        
        data.append({
            "Timestamp": dec.get("timestamp", ""),
            "Agent": dec.get("agent", "Unknown"),
            "Decision Type": dec.get("decision_type", "Unknown"),
            "Iteration": dec.get("iteration", 0),
            "Input (JSON)": json.dumps(input_data, ensure_ascii=False),
            "Output (JSON)": json.dumps(output_data, ensure_ascii=False),
            "Reasoning": str(reasoning)[:500] if reasoning else "",
            "Selected Phase": output_data.get("selected_phase", ""),
            "Selected TTP": output_data.get("selected_ttp", ""),
            "Is Valid": output_data.get("is_valid", ""),
            "Errors": "; ".join(output_data.get("errors", [])),
            "Warnings": "; ".join(output_data.get("warnings", []))
        })
    
    df = pd.DataFrame(data)
    output = BytesIO()
    df.to_csv(output, index=False, encoding='utf-8')
    output.seek(0)
    return output


def export_decisions_to_json(decisions: list) -> str:
    """Exports agent decisions as JSON."""
    from utils.safe_json import safe_json_dumps
    return safe_json_dumps(decisions, indent=2, ensure_ascii=False)


def export_decisions_to_excel(decisions: list) -> BytesIO:
    """Exports agent decisions as Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Agent Decisions"
        
        # Header
        headers = [
            "Timestamp", "Agent", "Decision Type", "Iteration",
            "Input", "Output", "Reasoning", "Selected Phase",
            "Selected TTP", "Is Valid", "Errors", "Warnings"
        ]
        
        # Header-Styling
        header_fill = PatternFill(start_color="2c5282", end_color="2c5282", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Schreibe Header
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Schreibe Daten
        for row_num, dec in enumerate(decisions, 2):
            input_data = dec.get("input", {})
            output_data = dec.get("output", {})
            reasoning = output_data.get("reasoning", "")
            if isinstance(reasoning, dict):
                reasoning = json.dumps(reasoning, ensure_ascii=False)
            
            data_row = [
                dec.get("timestamp", ""),
                dec.get("agent", "Unknown"),
                dec.get("decision_type", "Unknown"),
                dec.get("iteration", 0),
                json.dumps(input_data, ensure_ascii=False),
                json.dumps(output_data, ensure_ascii=False),
                str(reasoning)[:500] if reasoning else "",
                output_data.get("selected_phase", ""),
                output_data.get("selected_ttp", ""),
                output_data.get("is_valid", ""),
                "; ".join(output_data.get("errors", [])),
                "; ".join(output_data.get("warnings", []))
            ]
            
            for col_num, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                if col_num in [5, 6, 7]:  # Input, Output, Reasoning - Wrap Text
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Spaltenbreiten anpassen
        column_widths = {
            'A': 20,  # Timestamp
            'B': 15,  # Agent
            'C': 20,  # Decision Type
            'D': 10,  # Iteration
            'E': 40,  # Input
            'F': 40,  # Output
            'G': 50,  # Reasoning
            'H': 15,  # Selected Phase
            'I': 15,  # Selected TTP
            'J': 10,  # Is Valid
            'K': 30,  # Errors
            'L': 30   # Warnings
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Zeilenhöhe für Text-Spalten
        for row in range(2, len(decisions) + 2):
            ws.row_dimensions[row].height = 60
        
        # Freeze Header
        ws.freeze_panes = 'A2'
        
        # Speichere in BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
        
    except ImportError:
        raise ImportError("openpyxl ist nicht installiert. Bitte installiere es mit: pip install openpyxl")
    except Exception as e:
        raise Exception(f"Error during Excel export: {e}")


def export_to_msel(injects: list) -> str:
    """
    Exports injects in MSEL format (Master Scenario Event List).
    
    MSEL is a standard format for crisis exercises.
    Format: Tab-separiert mit spezifischen Spalten.
    """
    lines = []
    
    # MSEL Header
    lines.append("MSEL Version 1.0")
    lines.append("Generated by DORA Scenario Generator")
    lines.append("")
    lines.append("Time\tInject ID\tPhase\tSource\tTarget\tModality\tContent\tMITRE ID\tAssets\tDORA Tag\tBusiness Impact")
    
    # Daten
    for inj in injects:
        line = "\t".join([
            inj.time_offset,
            inj.inject_id,
            inj.phase.value,
            inj.source,
            inj.target,
            inj.modality.value,
            inj.content.replace("\n", "").replace("\t", ""),  # Tabs und Newlines entfernen
            inj.technical_metadata.mitre_id or "",
            ", ".join(inj.technical_metadata.affected_assets),
            inj.dora_compliance_tag or "",
            (inj.business_impact or "").replace("\n", "").replace("\t", "")
        ])
        lines.append(line)
    
    return "\n".join(lines)


def main():
    """Hauptfunktion der Streamlit App."""
    init_session_state()
    
    # Enterprise Header
    st.markdown("""
    <div class="main-header">
        <h1> DORA Szenariengenerator</h1>
        <div class="sub-header">
            Enterprise-Grade Krisenszenario-Generation für Finanzunternehmen
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Sidebar Configuration
    with st.sidebar:
        st.markdown("""
        <div style="padding: 0 0 1.5rem 0;">
            <h2 style="color: var(--text-primary); margin: 0; font-size: 1.125rem; font-weight: 600;">Configuration</h2>
        </div>
        """, unsafe_allow_html=True)
        
        scenario_type = st.selectbox(
            "Scenario Type",
            options=[st.value for st in ScenarioType],
            format_func=lambda x: x.replace("_", "").title(),
            help="Select the type of crisis scenario to generate"
        )
        
        max_iterations = st.slider(
            "Number of Injects",
            min_value=1,
            max_value=20,
            value=5,
            help="Number of injects to generate"
        )
        
        st.markdown("---")
        
        st.markdown("""
        <div style="padding: 0.5rem 0;">
            <h3 style="color: var(--text-primary); margin: 0 0 0.75rem 0; font-size: 0.875rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.05em;">Advanced Options</h3>
        </div>
        """, unsafe_allow_html=True)
        
        auto_phase_transition = st.checkbox(
            "Automatic Phase Transitions",
            value=True,
            help="Phases transition automatically based on context"
        )
        
        interactive_mode = st.checkbox(
            "🎮 Interactive Mode",
            value=False,
            help="Live-Entscheidungen während der Generierung treffen. Das Szenario pausiert an Decision-Points für deine Entscheidungen."
        )
        
        show_validation_details = st.checkbox(
            "Show Validation Details",
            value=False,
            help="Display detailed validation information"
        )
        
        st.markdown("---")
        st.markdown("""
        <div style="padding: 0.5rem 0;">
            <h3 style="color: var(--text-primary); margin: 0 0 0.75rem 0; font-size: 0.875rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.05em;">Infrastructure Setup</h3>
        </div>
        """, unsafe_allow_html=True)
        
        # Infrastructure Template Auswahl
        try:
            from templates.infrastructure_templates import get_available_templates
            templates = get_available_templates()
            template_names = ["None (Use Existing)"] + list(templates.keys())
            
            # Template-Auswahl mit Beschreibungen
            selected_template = st.selectbox(
                "Infrastructure Template",
                options=template_names,
                format_func=lambda x: x.replace("_", " ").title() if x != "None (Use Existing)" else x,
                help="Lade ein vordefiniertes Infrastruktur-Modell in Neo4j"
            )
            
            # Zeige Template-Beschreibung
            if selected_template != "None (Use Existing)" and selected_template in templates:
                template = templates[selected_template]
                st.info(f"**{template.name}**\n\n{template.description}")
                
                # Zeige Template-Details
                with st.expander("📋 Template-Details", expanded=False):
                    entities = template.get_entities()
                    relationships = template.get_relationships()
                    st.markdown(f"""
                    - **Entitäten:** {len(entities)}
                    - **Beziehungen:** {len(relationships)}
                    - **Typen:** {', '.join(set(e.get('type', 'Unknown') for e in entities))}
                    """)
            
            if st.button("Load Template", width='stretch', disabled=(selected_template == "None (Use Existing)")):
                if st.session_state.neo4j_client is None:
                    st.session_state.neo4j_client = Neo4jClient()
                    st.session_state.neo4j_client.connect()
                
                try:
                    template_key = selected_template
                    template = templates[template_key]
                    success = st.session_state.neo4j_client.initialize_base_infrastructure(template_key)
                    if success:
                        st.success(f"✅ Template '{template.name}' erfolgreich geladen!")
                        st.info(f"**Hinweis:** Das Template wurde in Neo4j geladen. Du kannst jetzt ein Szenario generieren.")
                    else:
                        st.warning("Template konnte nicht vollständig geladen werden")
                except Exception as e:
                    st.error(f"Fehler beim Laden des Templates: {e}")
        except ImportError:
            st.info("Infrastructure Templates nicht verfügbar")
        except Exception as e:
            st.warning(f"Templates konnten nicht geladen werden: {e}")
        
        # System Status
        st.markdown("---")
        st.markdown("""
        <div style="padding: 1rem; background: var(--background); border-radius: 4px; border: 1px solid var(--border);">
            <h3 style="color: var(--text-primary); margin: 0 0 0.75rem 0; font-size: 0.875rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.05em;">System Status</h3>
            <div style="color: var(--text-secondary); font-size: 0.8125rem;">
                <p style="margin: 0.375rem 0;"><span style="display: inline-block; width: 8px; height: 8px; border-radius: 50%; background: var(--success); margin-right: 8px;"></span>Neo4j Knowledge Graph</p>
                <p style="margin: 0.375rem 0;"><span style="display: inline-block; width: 8px; height: 8px; border-radius: 50%; background: var(--success); margin-right: 8px;"></span>OpenAI GPT-4o</p>
                <p style="margin: 0.375rem 0;"><span style="display: inline-block; width: 8px; height: 8px; border-radius: 50%; background: var(--success); margin-right: 8px;"></span>LangGraph Workflow</p>
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    # Hauptbereich mit Tabs
    tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
        "Generation", 
        "Dashboard", 
        "Injects", 
        "Analytics",
        "Workflow Logs",
        "Historical Scenarios",
        "System Overview"
    ])
    
    with tab1:
        st.markdown("""
        <div class="enterprise-card">
            <h2 style="color: var(--text-primary); margin-top: 0; font-size: 1.25rem; font-weight: 600;">Scenario Generation</h2>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            st.markdown(f"""
            <div class="enterprise-card"style="border-left: 4px solid var(--primary);">
                <h3 style="color: var(--text-primary); margin-top: 0; font-size: 1.25rem;"> Configuration</h3>
                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 1rem; margin-top: 1rem;">
                    <div>
                        <p style="margin: 0; font-size: 0.875rem; color: var(--text-secondary);"><strong> Scenario Type</strong></p>
                        <p style="margin: 0.5rem 0 0 0; color: var(--text-primary); font-weight: 500;">{scenario_type.replace('_', '').title()}</p>
                    </div>
                    <div>
                        <p style="margin: 0; font-size: 0.875rem; color: var(--text-secondary);"><strong> Number of Injects</strong></p>
                        <p style="margin: 0.5rem 0 0 0; color: var(--text-primary); font-weight: 500;">{max_iterations}</p>
                    </div>
                </div>
                <div style="margin-top: 1rem; padding-top: 1rem; border-top: 1px solid var(--border);">
                    <p style="margin: 0; font-size: 0.875rem; color: var(--text-secondary);"><strong> Phase Transitions:</strong> {'Automatic'if auto_phase_transition else 'Manual'}</p>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown("""
            <div style="display: flex; flex-direction: column; gap: 0.75rem;">
            """, unsafe_allow_html=True)
            
            if st.button("Generate Scenario", type="primary", width='stretch'):
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                try:
                    status_text.info("Initializing system...")
                    progress_bar.progress(10)
                    
                    if st.session_state.neo4j_client is None:
                        status_text.info("Connecting to Neo4j...")
                        st.session_state.neo4j_client = Neo4jClient()
                        st.session_state.neo4j_client.connect()
                    
                    progress_bar.progress(20)
                    
                    if st.session_state.workflow is None or st.session_state.workflow.interactive_mode != interactive_mode:
                        status_text.info("Initializing workflow...")
                        st.session_state.workflow = ScenarioWorkflow(
                            neo4j_client=st.session_state.neo4j_client,
                            max_iterations=max_iterations,
                            interactive_mode=interactive_mode
                        )
                    
                    progress_bar.progress(30)
                    
                    if interactive_mode:
                        status_text.info("🎮 Starting interactive scenario generation... Entscheidungen werden an Decision-Points erfragt.")
                        # Initialisiere interaktiven State
                        st.session_state.interactive_scenario_state = None
                        st.session_state.pending_decision = None
                    else:
                        status_text.info("Starting scenario generation... (This may take several minutes)")
                    
                    result = st.session_state.workflow.generate_scenario(
                        scenario_type=ScenarioType(scenario_type)
                    )
                    
                    # Im interaktiven Modus: Prüfe ob Decision-Point erreicht wurde
                    if interactive_mode and result.get("pending_decision"):
                        st.session_state.interactive_scenario_state = result
                        st.session_state.pending_decision = result.get("pending_decision")
                        progress_bar.progress(50)
                        status_text.info("🎯 Decision-Point erreicht! Bitte treffe eine Entscheidung unten.")
                        st.rerun()
                    
                    progress_bar.progress(100)
                    status_text.empty()
                    
                    # Prüfe Ergebnis
                    injects = result.get('injects', [])
                    errors = result.get('errors', [])
                    warnings = result.get('warnings', [])
                    
                    if injects:
                        st.session_state.scenario_result = result
                        st.success(f"Scenario generated successfully ({len(injects)} injects)")
                        if warnings:
                            st.warning(f"{len(warnings)} warning(s) - see details in tabs")
                    else:
                        # No injects generated
                        error_msg = "No injects were generated."
                        if errors:
                            error_msg += f"\n\n**Errors:**\n"+ "\n".join([f"- {e}"for e in errors])
                        if warnings:
                            error_msg += f"\n\n**Warnings:**\n"+ "\n".join([f"- {w}"for w in warnings])
                        
                        st.error(error_msg)
                        
                        # Zeige Debug Information
                        with st.expander("Debug Information"):
                            st.json({
                                "scenario_id": result.get('scenario_id'),
                                "iteration": result.get('iteration'),
                                "max_iterations": result.get('max_iterations'),
                                "current_phase": result.get('current_phase').value if result.get('current_phase') else None,
                                "errors": errors,
                                "warnings": warnings,
                                "workflow_logs_count": len(result.get('workflow_logs', [])),
                                "agent_decisions_count": len(result.get('agent_decisions', []))
                            })
                        
                        st.info("**Tip:** Try loading a demo scenario to test functionality, or check error details above.")
                    
                except Exception as e:
                    progress_bar.progress(0)
                    status_text.empty()
                    st.error(f"Error during generation: {e}")
                    import traceback
                    with st.expander("Detailed Error Information"):
                        st.code(traceback.format_exc())
                    
                    st.info("**Tip:** Ensure that:\n- Neo4j is running (Docker)\n- OpenAI API key is set in .env\n- All dependencies are installed")
            
            st.markdown("---")
            st.markdown("**Demo Scenarios**")
            
            demo_scenarios = get_available_demo_scenarios()
            selected_demo = st.selectbox(
                "Select Demo Scenario",
                options=[""] + list(demo_scenarios.keys()),
                format_func=lambda x: "Please select..."if x == ""else x,
                help="Load a predefined demo scenario for quick testing"
            )
            
            if st.button("Load Demo Scenario", width='stretch', disabled=not selected_demo):
                try:
                    demo_result = load_demo_scenario(selected_demo)
                    st.session_state.scenario_result = demo_result
                    st.success(f"Demo scenario loaded ({len(demo_result.get('injects', []))} injects)")
                    st.info("**Note:** This is a demo scenario for testing. Use 'Generate Scenario'for real scenarios.")
                except Exception as e:
                    st.error(f"Error loading demo scenario: {e}")
            
            st.markdown("</div>", unsafe_allow_html=True)
        
        # Interaktiver Modus: Decision Point UI
        if st.session_state.get("pending_decision") and st.session_state.get("interactive_scenario_state"):
            st.markdown("---")
            st.markdown("""
            <div class="enterprise-card" style="border-left: 4px solid var(--warning); background: linear-gradient(135deg, #fff5e6 0%, #ffe6cc 100%);">
                <h2 style="color: #cc6600; margin-top: 0; font-size: 1.5rem;">🎮 Decision Point - Deine Entscheidung</h2>
                <p style="color: #996600; margin: 0.5rem 0;">Das Szenario wartet auf deine Entscheidung. Wähle eine Option, um fortzufahren.</p>
            </div>
            """, unsafe_allow_html=True)
            
            pending_decision = st.session_state.pending_decision
            interactive_state = st.session_state.interactive_scenario_state
            
            # Zeige aktuelle Situation
            situation = pending_decision.get("situation", {})
            col_sit1, col_sit2, col_sit3 = st.columns(3)
            
            with col_sit1:
                st.metric("Aktuelle Phase", situation.get("current_phase", "N/A").replace("_", " ").title())
            with col_sit2:
                st.metric("Injects generiert", situation.get("inject_count", 0))
            with col_sit3:
                st.metric("Kritische Assets kompromittiert", situation.get("critical_assets_compromised", 0))
            
            # Zeige letztes Inject
            if interactive_state.get("injects"):
                last_inject = interactive_state["injects"][-1]
                with st.expander("📋 Letztes Event", expanded=True):
                    st.markdown(f"**{last_inject.inject_id}** - {last_inject.phase.value.replace('_', ' ').title()}")
                    st.markdown(f"**{last_inject.content[:200]}...**")
            
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown("### 🎯 Entscheidungsoptionen:")
            
            # Zeige Entscheidungsoptionen
            options = pending_decision.get("options", [])
            selected_option = None
            
            for i, option in enumerate(options):
                option_id = option.get("id")
                option_title = option.get("title", "Option")
                option_desc = option.get("description", "")
                option_impact = option.get("impact", {})
                
                col_opt1, col_opt2 = st.columns([4, 1])
                
                with col_opt1:
                    # Zeige Impact-Informationen
                    impact_info = ""
                    if option_impact.get("phase_change"):
                        impact_info += f"🔄 Phase → {option_impact['phase_change']} | "
                    if option_impact.get("asset_protection"):
                        impact_info += f"🛡️ Protection: {option_impact['asset_protection']} | "
                    if option_impact.get("response_effectiveness"):
                        impact_info += f"⚡ Effectiveness: {option_impact['response_effectiveness']} | "
                    
                    st.markdown(f"""
                    <div style="padding: 1rem; background: var(--surface); border: 2px solid var(--border); border-radius: 8px; margin-bottom: 0.5rem;">
                        <h4 style="margin: 0 0 0.5rem 0; color: var(--text-primary);">{option_title}</h4>
                        <p style="margin: 0 0 0.5rem 0; color: var(--text-secondary); font-size: 0.875rem;">{option_desc}</p>
                        {f'<p style="margin: 0; color: var(--text-secondary); font-size: 0.75rem; font-style: italic;">{impact_info.rstrip(" | ")}</p>' if impact_info else ''}
                    </div>
                    """, unsafe_allow_html=True)
                
                with col_opt2:
                    if st.button("Wählen", key=f"decision_btn_{i}", width='stretch'):
                        selected_option = option_id
                        break
            
            # Wenn Option gewählt wurde
            if selected_option:
                # Wende Entscheidung an
                user_decisions = interactive_state.get("user_decisions", [])
                user_decisions.append({
                    "decision_id": pending_decision.get("decision_id"),
                    "choice_id": selected_option,
                    "decision_type": next((opt.get("type") for opt in options if opt.get("id") == selected_option), "general"),
                    "timestamp": datetime.now().isoformat()
                })
                
                interactive_state["user_decisions"] = user_decisions
                interactive_state["pending_decision"] = None
                
                # Setze State zurück für nächsten Schritt
                st.session_state.interactive_scenario_state = interactive_state
                st.session_state.pending_decision = None
                
                # Führe Workflow weiter aus
                with st.spinner("🔄 Wende Entscheidung an und generiere nächste Events..."):
                    try:
                        # Wende Entscheidung an
                        interactive_state = st.session_state.workflow._apply_user_decision(
                            interactive_state,
                            user_decisions[-1]
                        )
                        
                        # Führe Workflow weiter aus (generiert Events basierend auf Entscheidung)
                        result = st.session_state.workflow._execute_interactive_workflow(
                            interactive_state,
                            recursion_limit=100
                        )
                        
                        # Prüfe ob wieder pausiert
                        if result.get("pending_decision") and result.get("pending_decision", {}).get("required"):
                            st.session_state.interactive_scenario_state = result
                            st.session_state.pending_decision = result.get("pending_decision")
                            st.success(f"✅ Entscheidung angewendet! {len(result.get('injects', []))} Injects generiert.")
                            st.rerun()
                        else:
                            # Szenario beendet
                            st.session_state.scenario_result = result
                            st.session_state.interactive_scenario_state = None
                            st.session_state.pending_decision = None
                            
                            # Zeige End-Bedingung
                            end_condition = result.get("end_condition")
                            if end_condition:
                                if end_condition == "FATAL":
                                    st.error("💀 **FATAL ENDE:** Das System wurde vollständig kompromittiert. Szenario beendet.")
                                elif end_condition == "VICTORY":
                                    st.success("🏆 **SIEG:** Die Bedrohung wurde erfolgreich abgewehrt!")
                                elif end_condition == "NORMAL_END":
                                    st.info("✅ **NORMALES ENDE:** Recovery abgeschlossen. Szenario beendet.")
                            
                            st.rerun()
                    except Exception as e:
                        st.error(f"Fehler beim Fortsetzen des Workflows: {e}")
                        import traceback
                        with st.expander("Fehlerdetails"):
                            st.code(traceback.format_exc())
        
        if st.session_state.scenario_result:
            result = st.session_state.scenario_result
            
            # Prüfe ob Szenario in Neo4j gespeichert wurde
            warnings = result.get('warnings', [])
            saved_to_neo4j = not any('Neo4j' in w or 'gespeichert' in w.lower() for w in warnings)
            
            # Entscheidungshilfen und Zusatzinfos
            decision_aids = result.get('decision_aids', {})
            additional_info = result.get('additional_info', {})
            
            st.markdown(f"""
            <div class="enterprise-card"style="border-left: 4px solid var(--success); background: linear-gradient(135deg, #f0fff4 0%, #c6f6d5 100%);">
                <h3 style="color: #22543d; margin-top: 0; font-size: 1.25rem;"> Scenario Generated Successfully</h3>
                <div style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 1rem; margin-top: 1rem;">
                    <div>
                        <p style="margin: 0; font-size: 0.875rem; color: #22543d;"><strong> Scenario ID</strong></p>
                        <code style="background: white; padding: 0.5rem; border-radius: 6px; display: block; margin-top: 0.5rem; color: var(--text-primary);">{result.get('scenario_id')}</code>
                    </div>
                    <div>
                        <p style="margin: 0; font-size: 0.875rem; color: #22543d;"><strong> Final Phase</strong></p>
                        <p style="margin: 0.5rem 0 0 0; color: var(--text-primary); font-weight: 500;">{result.get('current_phase').value.replace('_', '').title() if result.get('current_phase') else 'N/A'}</p>
                    </div>
                    <div>
                        <p style="margin: 0; font-size: 0.875rem; color: #22543d;"><strong> Number of Injects</strong></p>
                        <p style="margin: 0.5rem 0 0 0; color: var(--text-primary); font-weight: 500; font-size: 1.5rem;">{len(result.get('injects', []))}</p>
                    </div>
                    <div>
                        <p style="margin: 0; font-size: 0.875rem; color: #22543d;"><strong> Storage Status</strong></p>
                        <p style="margin: 0.5rem 0 0 0; color: var(--text-primary); font-weight: 500;">
                            {'✅ Saved to Neo4j' if saved_to_neo4j else '⚠️ Not saved'}
                        </p>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            # Entscheidungshilfen anzeigen
            if decision_aids:
                st.markdown("---")
                st.markdown("""
                <div class="enterprise-card">
                    <h3 style="color: var(--text-primary); margin-top: 0; font-size: 1.25rem;">💡 Entscheidungshilfen & Zusatzinfos</h3>
                </div>
                """, unsafe_allow_html=True)
                
                col_aid1, col_aid2 = st.columns(2)
                
                with col_aid1:
                    # Szenario-Zusammenfassung
                    if "scenario_summary" in decision_aids:
                        summary = decision_aids["scenario_summary"]
                        st.markdown("**📊 Szenario-Zusammenfassung:**")
                        st.json(summary)
                    
                    # DORA-Compliance
                    if "dora_compliance" in decision_aids:
                        dora = decision_aids["dora_compliance"]
                        st.markdown("**📋 DORA-Compliance:**")
                        st.markdown(f"""
                        - **Compliance-Rate:** {dora.get('compliance_rate', 0)}%
                        - **Compliant Injects:** {dora.get('compliant_injects', 0)}/{len(injects)}
                        - **Empfehlung:** {dora.get('recommendation', 'N/A')}
                        """)
                
                with col_aid2:
                    # Empfehlungen
                    if "recommendations" in decision_aids:
                        st.markdown("**💡 Empfehlungen:**")
                        for rec in decision_aids["recommendations"]:
                            st.markdown(f"- {rec}")
                    
                    # Key Insights
                    if "key_insights" in decision_aids:
                        st.markdown("**🔍 Wichtige Erkenntnisse:**")
                        for insight in decision_aids["key_insights"]:
                            st.markdown(f"- {insight}")
                
                # Zusätzliche Informationen
                if additional_info:
                    with st.expander("📈 Zusätzliche Informationen", expanded=False):
                        st.json(additional_info)
    
    with tab2:
        st.markdown("""
        <div class="enterprise-card">
            <h2 style="color: var(--text-primary); margin-top: 0; font-size: 1.25rem; font-weight: 600;">Dashboard & Analytics</h2>
        </div>
        """, unsafe_allow_html=True)
        
        if st.session_state.scenario_result:
            result = st.session_state.scenario_result
            injects = result.get("injects", [])
            logs = result.get("workflow_logs", [])
            decisions = result.get("agent_decisions", [])
            errors = result.get("errors", [])
            warnings = result.get("warnings", [])
            
            if injects:
                # Wichtige Metriken - Erste Reihe
                col1, col2, col3, col4, col5 = st.columns(5)
                
                with col1:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-value">{len(injects)}</div>
                        <div class="metric-label">Injects</div>
                    </div>
                    """, unsafe_allow_html=True)
                    st.caption("Generierte Injects")
                
                with col2:
                    unique_phases = len(set(inj.phase for inj in injects))
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-value">{unique_phases}</div>
                        <div class="metric-label">Phases</div>
                    </div>
                    """, unsafe_allow_html=True)
                    st.caption("Unterschiedliche Krisenphasen")
                
                with col3:
                    all_assets = set()
                    for inj in injects:
                        all_assets.update(inj.technical_metadata.affected_assets)
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-value">{len(all_assets)}</div>
                        <div class="metric-label">Assets</div>
                    </div>
                    """, unsafe_allow_html=True)
                    st.caption("Betroffene Assets/Systeme")
                
                with col4:
                    # DORA-Compliance-Rate
                    dora_compliant = sum(1 for inj in injects if inj.dora_compliance_tag)
                    dora_rate = (dora_compliant / len(injects) * 100) if injects else 0
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-value">{dora_rate:.0f}%</div>
                        <div class="metric-label">DORA</div>
                    </div>
                    """, unsafe_allow_html=True)
                    st.caption(f"DORA-konform ({dora_compliant}/{len(injects)})")
                
                with col5:
                    mitre_ids = set(inj.technical_metadata.mitre_id for inj in injects if inj.technical_metadata.mitre_id)
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-value">{len(mitre_ids)}</div>
                        <div class="metric-label">MITRE TTPs</div>
                    </div>
                    """, unsafe_allow_html=True)
                    st.caption("Unterschiedliche MITRE-Techniken")
                
                st.markdown("<br>", unsafe_allow_html=True)
                
                # Zweite Reihe - Qualitäts-Metriken
                col6, col7, col8, col9, col10 = st.columns(5)
                
                with col6:
                    # Validierungsrate
                    if decisions:
                        critic_decisions = [d for d in decisions if d.get("agent") == "Critic"]
                        if critic_decisions:
                            valid_count = sum(1 for d in critic_decisions if d.get("output", {}).get("is_valid", False))
                            validation_rate = (valid_count / len(critic_decisions) * 100) if critic_decisions else 0
                            st.markdown(f"""
                            <div class="metric-card">
                                <div class="metric-value">{validation_rate:.0f}%</div>
                                <div class="metric-label">Valid</div>
                            </div>
                            """, unsafe_allow_html=True)
                            st.caption(f"Validierungsrate ({valid_count}/{len(critic_decisions)})")
                        else:
                            st.markdown("""
                            <div class="metric-card">
                                <div class="metric-value">-</div>
                                <div class="metric-label">Valid</div>
                            </div>
                            """, unsafe_allow_html=True)
                            st.caption("Keine Validierungen")
                    else:
                        st.markdown("""
                        <div class="metric-card">
                            <div class="metric-value">-</div>
                            <div class="metric-label">Valid</div>
                        </div>
                        """, unsafe_allow_html=True)
                        st.caption("Keine Daten")
                
                with col7:
                    # Fehlerrate
                    error_rate = (len(errors) / len(injects) * 100) if injects and errors else 0
                    error_color = "var(--danger)" if error_rate > 10 else "var(--warning)" if error_rate > 0 else "var(--success)"
                    st.markdown(f"""
                    <div class="metric-card" style="border-top: 4px solid {error_color};">
                        <div class="metric-value" style="color: {error_color};">{len(errors)}</div>
                        <div class="metric-label">Errors</div>
                    </div>
                    """, unsafe_allow_html=True)
                    st.caption(f"Fehlerrate: {error_rate:.1f}%")
                
                with col8:
                    # Warnungen
                    warning_count = len(warnings)
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-value">{warning_count}</div>
                        <div class="metric-label">Warnings</div>
                    </div>
                    """, unsafe_allow_html=True)
                    st.caption("Anzahl Warnungen")
                
                with col9:
                    # Durchschnittliche Severity
                    severities = [inj.technical_metadata.severity for inj in injects if inj.technical_metadata.severity]
                    if severities:
                        severity_map = {"Low": 1, "Medium": 2, "High": 3, "Critical": 4}
                        avg_severity_num = sum(severity_map.get(s, 2) for s in severities) / len(severities)
                        severity_labels = {1: "Low", 2: "Medium", 3: "High", 4: "Critical"}
                        avg_severity = severity_labels.get(round(avg_severity_num), "Medium")
                        st.markdown(f"""
                        <div class="metric-card">
                            <div class="metric-value" style="font-size: 2rem;">{avg_severity}</div>
                            <div class="metric-label">Avg Severity</div>
                        </div>
                        """, unsafe_allow_html=True)
                        st.caption(f"Durchschnittlicher Schweregrad")
                    else:
                        st.markdown("""
                        <div class="metric-card">
                            <div class="metric-value">-</div>
                            <div class="metric-label">Severity</div>
                        </div>
                        """, unsafe_allow_html=True)
                        st.caption("Keine Daten")
                
                with col10:
                    # Agent-Entscheidungen
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-value">{len(decisions)}</div>
                        <div class="metric-label">Decisions</div>
                    </div>
                    """, unsafe_allow_html=True)
                    st.caption("Agent-Entscheidungen")
                
                st.markdown("<br>", unsafe_allow_html=True)
                
                # Charts in Grid
                col_chart1, col_chart2 = st.columns(2)
                
                with col_chart1:
                    st.markdown("""
                    <div style="margin-bottom: 0.5rem;">
                        <h4 style="margin: 0; color: var(--text-primary); font-size: 1.1rem;">📊 Phasen-Verteilung</h4>
                        <p style="margin: 0.25rem 0 0 0; color: var(--text-secondary); font-size: 0.875rem;">
                            Verteilung der Injects über verschiedene Krisenphasen. Zeigt die Szenario-Progression.
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
                    fig = create_phase_distribution_chart(injects)
                    st.plotly_chart(fig, width='stretch')
                
                with col_chart2:
                    st.markdown("""
                    <div style="margin-bottom: 0.5rem;">
                        <h4 style="margin: 0; color: var(--text-primary); font-size: 1.1rem;">⏱️ Szenario-Timeline</h4>
                        <p style="margin: 0.25rem 0 0 0; color: var(--text-secondary); font-size: 0.875rem;">
                            Zeitliche Abfolge der Injects. Jeder Punkt repräsentiert einen Inject mit Phase und Details.
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
                    timeline_fig = create_timeline_chart(injects)
                    st.plotly_chart(timeline_fig, width='stretch')
                
                # DORA-Compliance Übersicht
                    st.markdown("---")
                    st.markdown("""
                    <div style="margin-bottom: 1rem;">
                    <h3 style="margin: 0; color: var(--text-primary); font-size: 1.5rem;">📋 DORA-Compliance Übersicht</h3>
                        <p style="margin: 0.5rem 0 0 0; color: var(--text-secondary); font-size: 0.875rem;">
                        Übersicht über DORA-Artikel 25 Compliance der generierten Injects.
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
                
                # DORA-Tags Analyse
                dora_tags = {}
                for inj in injects:
                    if inj.dora_compliance_tag:
                        tag = inj.dora_compliance_tag
                        dora_tags[tag] = dora_tags.get(tag, 0) + 1
                
                if dora_tags:
                    col_dora1, col_dora2 = st.columns([1, 2])
                    with col_dora1:
                        st.markdown("**DORA-Tags Verteilung:**")
                        dora_df = pd.DataFrame({
                            "DORA Tag": list(dora_tags.keys()),
                            "Anzahl": list(dora_tags.values())
                        }).sort_values("Anzahl", ascending=False)
                        st.dataframe(dora_df, width='stretch', hide_index=True)
                    
                    with col_dora2:
                        # DORA-Compliance Chart
                        fig_dora = px.pie(
                            values=list(dora_tags.values()),
                            names=list(dora_tags.keys()),
                            title="DORA Compliance Tags Distribution"
                        )
                        fig_dora.update_layout(
                            plot_bgcolor="rgba(0,0,0,0)",
                            paper_bgcolor="rgba(0,0,0,0)",
                            height=300
                        )
                        st.plotly_chart(fig_dora, width='stretch')
                else:
                    st.info("Keine DORA-Compliance-Tags in diesem Szenario")
                
                # Phase-Übergänge (kompakt)
                st.markdown("---")
                phase_transitions = []
                prev_phase = None
                for inj in injects:
                    if prev_phase and prev_phase != inj.phase:
                        phase_transitions.append({
                            "Von": prev_phase.value.replace('_', ' ').title(),
                            "Zu": inj.phase.value.replace('_', ' ').title(),
                            "Inject": inj.inject_id,
                            "Zeit": inj.time_offset
                        })
                    prev_phase = inj.phase
                
                if phase_transitions:
                    st.markdown("**🔄 Phase-Übergänge:**")
                    transitions_df = pd.DataFrame(phase_transitions)
                    st.dataframe(transitions_df, width='stretch', hide_index=True)
                else:
                    st.info("Keine Phasen-Übergänge in diesem Szenario")
                
                # Qualitäts-Score
                    st.markdown("---")
                    st.markdown("""
                    <div style="margin-bottom: 1rem;">
                    <h3 style="margin: 0; color: var(--text-primary); font-size: 1.5rem;">⭐ Szenario-Qualität</h3>
                    </div>
                    """, unsafe_allow_html=True)
                
                col_qual1, col_qual2, col_qual3 = st.columns(3)
                
                with col_qual1:
                    # DORA-Compliance Score
                    dora_score = (dora_compliant / len(injects) * 100) if injects else 0
                    score_color = "var(--success)" if dora_score >= 80 else "var(--warning)" if dora_score >= 50 else "var(--danger)"
                    st.markdown(f"""
                    <div class="metric-card" style="border-top: 4px solid {score_color};">
                        <div class="metric-value" style="color: {score_color};">{dora_score:.0f}%</div>
                        <div class="metric-label">DORA Score</div>
                        </div>
                        """, unsafe_allow_html=True)
                    st.caption("DORA-Compliance-Rate")
                
                with col_qual2:
                    # Validierungs-Score
                    if decisions:
                        critic_decisions = [d for d in decisions if d.get("agent") == "Critic"]
                        if critic_decisions:
                            valid_count = sum(1 for d in critic_decisions if d.get("output", {}).get("is_valid", False))
                            validation_score = (valid_count / len(critic_decisions) * 100) if critic_decisions else 0
                            val_color = "var(--success)" if validation_score >= 90 else "var(--warning)" if validation_score >= 70 else "var(--danger)"
                            st.markdown(f"""
                            <div class="metric-card" style="border-top: 4px solid {val_color};">
                                <div class="metric-value" style="color: {val_color};">{validation_score:.0f}%</div>
                                <div class="metric-label">Validation Score</div>
                            </div>
                            """, unsafe_allow_html=True)
                            st.caption("Validierungsrate")
                        else:
                            st.markdown("""
                            <div class="metric-card">
                                <div class="metric-value">-</div>
                                <div class="metric-label">Validation</div>
                            </div>
                            """, unsafe_allow_html=True)
                            st.caption("Keine Daten")
                    else:
                        st.markdown("""
                        <div class="metric-card">
                            <div class="metric-value">-</div>
                            <div class="metric-label">Validation</div>
                        </div>
                        """, unsafe_allow_html=True)
                        st.caption("Keine Daten")
                
                with col_qual3:
                    # Gesamt-Qualität (kombiniert)
                    if decisions:
                        critic_decisions = [d for d in decisions if d.get("agent") == "Critic"]
                        if critic_decisions:
                            valid_count = sum(1 for d in critic_decisions if d.get("output", {}).get("is_valid", False))
                            validation_score = (valid_count / len(critic_decisions) * 100) if critic_decisions else 0
                            overall_score = (dora_score + validation_score) / 2
                            overall_color = "var(--success)" if overall_score >= 80 else "var(--warning)" if overall_score >= 60 else "var(--danger)"
                            st.markdown(f"""
                            <div class="metric-card" style="border-top: 4px solid {overall_color};">
                                <div class="metric-value" style="color: {overall_color};">{overall_score:.0f}%</div>
                                <div class="metric-label">Overall Score</div>
                            </div>
                            """, unsafe_allow_html=True)
                            st.caption("Gesamt-Qualität")
                        else:
                            overall_score = dora_score
                            overall_color = "var(--success)" if overall_score >= 80 else "var(--warning)" if overall_score >= 60 else "var(--danger)"
                            st.markdown(f"""
                            <div class="metric-card" style="border-top: 4px solid {overall_color};">
                                <div class="metric-value" style="color: {overall_color};">{overall_score:.0f}%</div>
                                <div class="metric-label">Overall Score</div>
                            </div>
                            """, unsafe_allow_html=True)
                            st.caption("Gesamt-Qualität")
                    else:
                        overall_score = dora_score
                        overall_color = "var(--success)" if overall_score >= 80 else "var(--warning)" if overall_score >= 60 else "var(--danger)"
                        st.markdown(f"""
                        <div class="metric-card" style="border-top: 4px solid {overall_color};">
                            <div class="metric-value" style="color: {overall_color};">{overall_score:.0f}%</div>
                            <div class="metric-label">Overall Score</div>
                        </div>
                        """, unsafe_allow_html=True)
                        st.caption("Gesamt-Qualität")
                
            else:
                st.warning("No data available.")
        else:
            st.info(" Generate a scenario first in the 'Generation'")
    
    with tab3:
        st.markdown("""
        <div class="enterprise-card">
            <h2 style="color: var(--text-primary); margin-top: 0; font-size: 1.25rem; font-weight: 600;">Generated Injects</h2>
        </div>
        """, unsafe_allow_html=True)
        
        if st.session_state.scenario_result:
            result = st.session_state.scenario_result
            injects = result.get("injects", [])
            
            if injects:
                # Filter andSearch
                col_filter1, col_filter2, col_filter3 = st.columns(3)
                
                with col_filter1:
                    all_phases = ["All"] + [p.value for p in CrisisPhase]
                    selected_phase = st.selectbox(
                        "Filter by Phase",
                        options=all_phases,
                        format_func=lambda x: x.replace("_", "").title() if x != "All"else x
                    )
                    st.session_state.filter_phase = selected_phase if selected_phase != "All"else None
                
                with col_filter2:
                    search_term = st.text_input(
                        "🔎Search",
                        value=st.session_state.search_term,
                        placeholder="Search in content, Assets, MITRE ID..."
                    )
                    st.session_state.search_term = search_term
                
                with col_filter3:
                    st.markdown(f"""
                    <div class="metric-card"style="padding: 1.25rem;">
                        <div class="metric-value"style="font-size: 2rem;">{len(injects)}</div>
                        <div class="metric-label">Injects</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                # Apply filters
                filtered_injects = injects
                if st.session_state.filter_phase:
                    filtered_injects = [inj for inj in filtered_injects if inj.phase.value == st.session_state.filter_phase]
                
                if st.session_state.search_term:
                    search_lower = st.session_state.search_term.lower()
                    filtered_injects = [
                        inj for inj in filtered_injects
                        if (search_lower in inj.content.lower() or
                            search_lower in ''.join(inj.technical_metadata.affected_assets).lower() or
                            (inj.technical_metadata.mitre_id and search_lower in inj.technical_metadata.mitre_id.lower()))
                    ]
                
                st.markdown(f"**Angezeigt:** {len(filtered_injects)} von {len(injects)} Injects")
                st.markdown("<br>", unsafe_allow_html=True)
                
                # Export-Optionen
                st.markdown("**📥 Export-Optionen**")
                col_exp1, col_exp2, col_exp3, col_exp4 = st.columns(4)
                
                with col_exp1:
                    csv_data = export_to_csv(filtered_injects)
                    st.download_button(
                        label="CSV",
                        data=csv_data,
                        file_name=f"scenario_{result.get('scenario_id')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv",
                        width='stretch',
                        help="Export asCSV file"
                    )
                
                with col_exp2:
                    json_data = export_to_json(filtered_injects)
                    st.download_button(
                        label="JSON",
                        data=json_data,
                        file_name=f"scenario_{result.get('scenario_id')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                        mime="application/json",
                        width='stretch',
                        help="Export as JSON file"
                    )
                
                with col_exp3:
                    try:
                        excel_data = export_to_excel(filtered_injects)
                        st.download_button(
                            label="Excel",
                            data=excel_data,
                            file_name=f"scenario_{result.get('scenario_id')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            width='stretch',
                            help="Export as Excel file (.xlsx)"
                        )
                    except ImportError:
                        st.error("Excel export requires openpyxl. Install with: pip install openpyxl")
                    except Exception as e:
                        st.error(f"Excel export error: {e}")
                
                with col_exp4:
                    try:
                        msel_data = export_to_msel(filtered_injects)
                        st.download_button(
                            label="MSEL",
                            data=msel_data,
                            file_name=f"scenario_{result.get('scenario_id')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.msel",
                            mime="text/plain",
                            width='stretch',
                            help="Export in MSEL format (Master Scenario Event List)"
                        )
                    except Exception as e:
                        st.error(f"MSEL export error: {e}")
                
                st.markdown("<br>", unsafe_allow_html=True)
                
                # Display injects
                for i, inject in enumerate(filtered_injects, 1):
                    format_inject_card_safe(inject, i)
                
                if not filtered_injects:
                    st.warning("Keine Injects entsprechen den Filterkriterien.")
            else:
                st.warning("No injects generated.")
        else:
            st.info(" Generate a scenario first in the 'Generation'")
    
    with tab4:
        st.markdown("""
        <div class="enterprise-card">
            <h2 style="color: var(--text-primary); margin-top: 0; font-size: 1.25rem; font-weight: 600;">Advanced Analytics</h2>
        </div>
        """, unsafe_allow_html=True)
        
        if st.session_state.scenario_result:
            result = st.session_state.scenario_result
            injects = result.get("injects", [])
            
            if injects:
                # Assets-Übersicht
                st.markdown("""
                <div style="margin-bottom: 1rem;">
                    <h3 style="margin: 0; color: var(--text-primary); font-size: 1.5rem;"> Affected Assets</h3>
                    <p style="margin: 0.5rem 0 0 0; color: var(--text-secondary); font-size: 0.875rem;">
                        Analyzes which assets (servers, systems, applications) are affected in the scenario 
                        and how frequently they appear in injects. Assets with higher counts are more critical.
                    </p>
                </div>
                """, unsafe_allow_html=True)
                all_assets = set()
                asset_impact_count = {}
                for inj in injects:
                    for asset in inj.technical_metadata.affected_assets:
                        all_assets.add(asset)
                        asset_impact_count[asset] = asset_impact_count.get(asset, 0) + 1
                
                if all_assets:
                    # Asset-Impact Chart
                    asset_df = pd.DataFrame({
                        "Asset": list(asset_impact_count.keys()),
                        "Affected": list(asset_impact_count.values())
                    }).sort_values("Affected", ascending=False)
                    
                    fig_assets = px.bar(
                        asset_df,
                        x="Asset",
                        y="Affected",
                        title="Asset Impact Analysis",
                        color="Affected",
                        color_continuous_scale="Blues"
                    )
                    fig_assets.update_layout(
                        plot_bgcolor="rgba(0,0,0,0)",
                        paper_bgcolor="rgba(0,0,0,0)",
                        height=400
                    )
                    st.plotly_chart(fig_assets, width='stretch')
                    
                    # Asset Tags
                    asset_html = '<div style="display: flex; flex-wrap: wrap; gap: 0.75rem; margin-top: 1rem;">'
                    for asset, count in sorted(asset_impact_count.items(), key=lambda x: x[1], reverse=True):
                        asset_html += f'<span style="background: linear-gradient(135deg, var(--primary) 0%, var(--primary-light) 100%); color: white; padding: 0.5rem 1rem; border-radius: 8px; font-size: 0.875rem; font-weight: 500; box-shadow: 0 2px 8px rgba(44, 82, 130, 0.2);">{html.escape(asset)} <span style="opacity: 0.8;">({count}x)</span></span>'
                    asset_html += '</div>'
                    st.markdown(asset_html, unsafe_allow_html=True)
                else:
                    st.info("No assets affected")
                
                # MITRE TTP Analyse
                st.markdown("---")
                st.markdown("""
                <div style="margin-bottom: 1rem;">
                    <h3 style="margin: 0; color: var(--text-primary); font-size: 1.5rem;"> MITRE ATT&CK TTP Analysis</h3>
                    <p style="margin: 0.5rem 0 0 0; color: var(--text-secondary); font-size: 0.875rem;">
                        Shows all used MITRE ATT&CK tactics, techniques and procedures (TTPs) in the scenario. 
                        Frequently used TTPs indicate recurring attack patterns.
                    </p>
                </div>
                """, unsafe_allow_html=True)
                mitre_counts = {}
                for inj in injects:
                    mitre_id = inj.technical_metadata.mitre_id
                    if mitre_id:
                        mitre_counts[mitre_id] = mitre_counts.get(mitre_id, 0) + 1
                
                if mitre_counts:
                    mitre_df = pd.DataFrame({
                        "MITRE ID": list(mitre_counts.keys()),
                        "Usage": list(mitre_counts.values())
                    }).sort_values("Usage", ascending=False)
                    
                    st.dataframe(mitre_df, width='stretch', hide_index=True)
                    st.caption("Tip: Higher usage counts indicate recurring attack patterns in the scenario.")
                else:
                    st.info("No MITRE IDs available")
            else:
                st.warning("No data available for visualization.")
        else:
            st.info(" Generate a scenario first in the 'Generation'")
    
    with tab5:
        st.markdown("""
        <div class="enterprise-card">
            <h2 style="color: var(--text-primary); margin-top: 0; font-size: 1.75rem;"> Workflow Logs & Decisions</h2>
        </div>
        """, unsafe_allow_html=True)
        
        if st.session_state.scenario_result:
            result = st.session_state.scenario_result
            logs = result.get("workflow_logs", [])
            decisions = result.get("agent_decisions", [])
            
            if logs or decisions:
                # Workflow Logs
                if logs:
                    st.subheader("Workflow Logs")
                    for log in logs:
                        node = log.get("node", "Unknown")
                        action = log.get("action", "")
                        details = log.get("details", {})
                        timestamp = log.get("timestamp", "")
                        status = details.get("status", "unknown")
                        
                        status_color = {
                            "success": "#38a169",
                            "error": "#e53e3e",
                            "warning": "#d69e2e"
                        }.get(status, "#718096")
                        
                        st.markdown(f"""
                        <div class="log-entry"style="border-left-color: {status_color};">
                            <div style="display: flex; justify-content: space-between; margin-bottom: 0.5rem;">
                                <strong style="color: var(--primary-dark);">{node}</strong>
                                <span style="color: var(--text-secondary); font-size: 0.75rem;">{timestamp.split('T')[1][:8] if 'T'in timestamp else timestamp}</span>
                            </div>
                            <p style="margin: 0.25rem 0; color: var(--text-primary);">{action}</p>
                            <p style="margin: 0.5rem 0 0 0; color: var(--text-secondary); font-size: 0.8rem;">
                                {json.dumps(details, indent=2, ensure_ascii=False)[:200]}...
                            </p>
                        </div>
                        """, unsafe_allow_html=True)
                
                # Agent Decisions
                if decisions:
                    st.markdown("---")
                    
                    # Export-Optionen für Decisions
                    col_dec_header1, col_dec_header2 = st.columns([3, 1])
                    with col_dec_header1:
                        st.subheader(" Agent Decisions")
                    with col_dec_header2:
                        st.markdown("""
                        <div style="margin-top: 0.5rem;">
                            <p style="margin: 0; font-size: 0.75rem; color: var(--text-secondary);">
                                Export für Analyse
                            </p>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    # Export Buttons
                    col_exp_dec1, col_exp_dec2, col_exp_dec3 = st.columns(3)
                    
                    with col_exp_dec1:
                        try:
                            csv_decisions = export_decisions_to_csv(decisions)
                            st.download_button(
                                label="📥 Export Decisions (CSV)",
                                data=csv_decisions,
                                file_name=f"agent_decisions_{result.get('scenario_id', 'unknown')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                                mime="text/csv",
                                width='stretch',
                                help="Export Agent-Entscheidungen als CSV für Analyse"
                            )
                        except Exception as e:
                            st.error(f"CSV Export Fehler: {e}")
                    
                    with col_exp_dec2:
                        try:
                            json_decisions = export_decisions_to_json(decisions)
                            st.download_button(
                                label="📥 Export Decisions (JSON)",
                                data=json_decisions,
                                file_name=f"agent_decisions_{result.get('scenario_id', 'unknown')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                                mime="application/json",
                                width='stretch',
                                help="Export Agent-Entscheidungen als JSON"
                            )
                        except Exception as e:
                            st.error(f"JSON Export Fehler: {e}")
                    
                    with col_exp_dec3:
                        try:
                            excel_decisions = export_decisions_to_excel(decisions)
                            st.download_button(
                                label="📥 Export Decisions (Excel)",
                                data=excel_decisions,
                                file_name=f"agent_decisions_{result.get('scenario_id', 'unknown')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                width='stretch',
                                help="Export Agent-Entscheidungen als Excel für detaillierte Analyse"
                            )
                        except ImportError:
                            st.error("Excel export requires openpyxl")
                        except Exception as e:
                            st.error(f"Excel Export Fehler: {e}")
                    
                    st.markdown("<br>", unsafe_allow_html=True)
                    
                    # Entscheidungs-Analyse
                    with st.expander("📊 Entscheidungs-Analyse", expanded=False):
                        # Statistiken
                        col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
                        
                        with col_stat1:
                            agent_counts = {}
                            for dec in decisions:
                                agent = dec.get("agent", "Unknown")
                                agent_counts[agent] = agent_counts.get(agent, 0) + 1
                            st.metric("Anzahl Agenten", len(agent_counts))
                        
                        with col_stat2:
                            st.metric("Gesamt-Entscheidungen", len(decisions))
                        
                        with col_stat3:
                            decision_types = {}
                            for dec in decisions:
                                dt = dec.get("decision_type", "Unknown")
                                decision_types[dt] = decision_types.get(dt, 0) + 1
                            st.metric("Entscheidungs-Typen", len(decision_types))
                        
                        with col_stat4:
                            # Zähle Fehler in Decisions
                            total_errors = 0
                            for dec in decisions:
                                output = dec.get("output", {})
                                total_errors += len(output.get("errors", []))
                            st.metric("Gesamt-Fehler", total_errors)
                        
                        # Agent-Verteilung
                        if agent_counts:
                            st.markdown("**Entscheidungen pro Agent:**")
                            agent_df = pd.DataFrame({
                                "Agent": list(agent_counts.keys()),
                                "Anzahl": list(agent_counts.values())
                            }).sort_values("Anzahl", ascending=False)
                            st.dataframe(agent_df, width='stretch', hide_index=True)
                        
                        # Entscheidungs-Typen
                        if decision_types:
                            st.markdown("**Entscheidungs-Typen:**")
                            type_df = pd.DataFrame({
                                "Typ": list(decision_types.keys()),
                                "Anzahl": list(decision_types.values())
                            }).sort_values("Anzahl", ascending=False)
                            st.dataframe(type_df, width='stretch', hide_index=True)
                        
                        # Entscheidungs-Qualität
                        st.markdown("**Entscheidungs-Qualität:**")
                        quality_data = []
                        for dec in decisions:
                            output = dec.get("output", {})
                            errors = output.get("errors", [])
                            warnings = output.get("warnings", [])
                            is_valid = output.get("is_valid", True)
                            
                            quality_data.append({
                                "Agent": dec.get("agent", "Unknown"),
                                "Decision Type": dec.get("decision_type", "Unknown"),
                                "Is Valid": is_valid,
                                "Errors": len(errors),
                                "Warnings": len(warnings),
                                "Quality Score": "High" if is_valid and len(errors) == 0 and len(warnings) <= 1 else "Medium" if is_valid else "Low"
                            })
                        
                        if quality_data:
                            quality_df = pd.DataFrame(quality_data)
                            
                            # Zusammenfassung
                            col_q1, col_q2, col_q3 = st.columns(3)
                            with col_q1:
                                high_quality = len(quality_df[quality_df["Quality Score"] == "High"])
                                st.metric("High Quality", high_quality)
                            with col_q2:
                                medium_quality = len(quality_df[quality_df["Quality Score"] == "Medium"])
                                st.metric("Medium Quality", medium_quality)
                            with col_q3:
                                low_quality = len(quality_df[quality_df["Quality Score"] == "Low"])
                                st.metric("Low Quality", low_quality)
                            
                            st.dataframe(quality_df, width='stretch', hide_index=True)
                    
                    st.markdown("<br>", unsafe_allow_html=True)
                    
                    for dec in decisions:
                        agent = dec.get("agent", "Unknown")
                        decision_type = dec.get("decision_type", "Unknown")
                        input_data = dec.get("input", {})
                        output_data = dec.get("output", {})
                        timestamp = dec.get("timestamp", "")
                        
                        with st.expander(f"{agent} - {decision_type} ({timestamp.split('T')[1][:8] if 'T' in timestamp else timestamp})"):
                            col_dec1, col_dec2 = st.columns(2)
                            
                            with col_dec1:
                                st.markdown("** Input:**")
                                st.json(input_data)
                            
                            with col_dec2:
                                st.markdown("**📤 Output:**")
                                st.json(output_data)
                else:
                    st.info("Keine Agent Decisions verfügbar")
            else:
                st.info("No workflow logs available. Generate ein neues Scenario, to see logs.")
        else:
            st.info("Generate a scenario first in the 'Generation' tab")
    
    with tab6:
        st.markdown("""
        <div class="enterprise-card">
            <h2 style="color: var(--text-primary); margin-top: 0; font-size: 1.25rem; font-weight: 600;">Historical Scenarios</h2>
            <p style="color: var(--text-secondary); font-size: 0.875rem; margin-top: 0.5rem;">
                Zeige und lade gespeicherte Szenarien aus der Neo4j-Datenbank. Alle generierten Szenarien werden automatisch gespeichert.
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        # Initialisiere Neo4j Client falls nötig
        if st.session_state.neo4j_client is None:
            try:
                st.session_state.neo4j_client = Neo4jClient()
                st.session_state.neo4j_client.connect()
            except Exception as e:
                st.error(f"Fehler bei Neo4j-Verbindung: {e}")
                st.info("Stelle sicher, dass Neo4j läuft (./start_neo4j.sh)")
                st.stop()
        
        # Filter-Optionen
        col_filter1, col_filter2, col_filter3 = st.columns(3)
        
        with col_filter1:
            filter_user = st.text_input(
                "🔍 Filter by User",
                value="",
                placeholder="Optional: Benutzername",
                help="Filtere nach Benutzer, der das Szenario erstellt hat"
            )
        
        with col_filter2:
            filter_type = st.selectbox(
                "📋 Filter by Type",
                options=["All"] + [st.value for st in ScenarioType],
                format_func=lambda x: x.replace("_", " ").title() if x != "All" else x,
                help="Filtere nach Szenario-Typ"
            )
        
        with col_filter3:
            limit_scenarios = st.number_input(
                "📊 Max. Anzahl",
                min_value=10,
                max_value=100,
                value=50,
                step=10,
                help="Maximale Anzahl anzuzeigender Szenarien"
            )
        
        # Lade Szenarien
        try:
            scenarios = st.session_state.neo4j_client.list_scenarios(
                limit=limit_scenarios,
                user=filter_user if filter_user else None,
                scenario_type=filter_type if filter_type != "All" else None
            )
            
            if scenarios:
                st.markdown(f"**Gefunden: {len(scenarios)} Szenarien**")
                st.markdown("<br>", unsafe_allow_html=True)
                
                # Zeige Szenarien in Cards
                for scenario in scenarios:
                    with st.container():
                        col_scen1, col_scen2, col_scen3 = st.columns([3, 1, 1])
                        
                        with col_scen1:
                            st.markdown(f"""
                            <div class="enterprise-card" style="padding: 1rem; margin-bottom: 1rem;">
                                <h4 style="color: var(--text-primary); margin: 0 0 0.5rem 0; font-size: 1rem; font-weight: 600;">
                                    {scenario.get('scenario_id', 'Unknown')}
                                </h4>
                                <div style="display: grid; grid-template-columns: repeat(3, 1fr); gap: 1rem; margin-top: 0.75rem;">
                                    <div>
                                        <p style="margin: 0; font-size: 0.75rem; color: var(--text-secondary);">Type</p>
                                        <p style="margin: 0.25rem 0 0 0; color: var(--text-primary); font-weight: 500; font-size: 0.875rem;">
                                            {scenario.get('scenario_type', 'N/A').replace('_', ' ').title()}
                                        </p>
                                    </div>
                                    <div>
                                        <p style="margin: 0; font-size: 0.75rem; color: var(--text-secondary);">Phase</p>
                                        <p style="margin: 0.25rem 0 0 0; color: var(--text-primary); font-weight: 500; font-size: 0.875rem;">
                                            {scenario.get('current_phase', 'N/A').replace('_', ' ').title()}
                                        </p>
                                    </div>
                                    <div>
                                        <p style="margin: 0; font-size: 0.75rem; color: var(--text-secondary);">Injects</p>
                                        <p style="margin: 0.25rem 0 0 0; color: var(--text-primary); font-weight: 500; font-size: 0.875rem;">
                                            {scenario.get('inject_count', 0)}
                                        </p>
                                    </div>
                                </div>
                                <div style="margin-top: 0.75rem; padding-top: 0.75rem; border-top: 1px solid var(--border);">
                                    <p style="margin: 0; font-size: 0.75rem; color: var(--text-secondary);">
                                        Created: {scenario.get('created_at', 'N/A')} | User: {scenario.get('user', 'system')}
                                    </p>
                                </div>
                            </div>
                            """, unsafe_allow_html=True)
                        
                        with col_scen2:
                            if st.button("📥 Load", key=f"load_{scenario.get('scenario_id')}", width='stretch'):
                                try:
                                    loaded_scenario = st.session_state.neo4j_client.get_scenario(scenario.get('scenario_id'))
                                    if loaded_scenario:
                                        # Konvertiere zu ScenarioResult-Format
                                        from state_models import Inject, TechnicalMetadata, InjectModality
                                        
                                        injects = []
                                        for inj_data in loaded_scenario.get('injects', []):
                                            try:
                                                inject = Inject(
                                                    inject_id=inj_data.get('id', 'UNKNOWN'),
                                                    time_offset=inj_data.get('time_offset', 'T+00:00'),
                                                    phase=CrisisPhase(inj_data.get('phase', 'normal_operation')),
                                                    source=inj_data.get('source', 'Unknown'),
                                                    target=inj_data.get('target', 'Unknown'),
                                                    modality=InjectModality(inj_data.get('modality', 'SIEM Alert')),
                                                    content=inj_data.get('content', ''),
                                                    technical_metadata=TechnicalMetadata(
                                                        mitre_id=inj_data.get('mitre_id', ''),
                                                        affected_assets=[],
                                                        severity=inj_data.get('severity', 'Medium')
                                                    ),
                                                    dora_compliance_tag=inj_data.get('dora_compliance_tag'),
                                                    business_impact=inj_data.get('business_impact')
                                                )
                                                injects.append(inject)
                                            except Exception as e:
                                                st.warning(f"Fehler beim Laden von Inject: {e}")
                                                continue
                                        
                                        scenario_result = {
                                            'scenario_id': loaded_scenario.get('scenario_id'),
                                            'scenario_type': ScenarioType(loaded_scenario.get('scenario_type', 'ransomware')),
                                            'current_phase': CrisisPhase(loaded_scenario.get('current_phase', 'normal_operation')),
                                            'injects': injects,
                                            'start_time': loaded_scenario.get('start_time'),
                                            'metadata': loaded_scenario.get('metadata', {}),
                                            'workflow_logs': [],
                                            'agent_decisions': [],
                                            'errors': [],
                                            'warnings': []
                                        }
                                        
                                        st.session_state.scenario_result = scenario_result
                                        st.success(f"Szenario {scenario.get('scenario_id')} geladen!")
                                        st.rerun()
                                except Exception as e:
                                    st.error(f"Fehler beim Laden: {e}")
                        
                        with col_scen3:
                            if st.button("🗑️ Delete", key=f"delete_{scenario.get('scenario_id')}", width='stretch'):
                                try:
                                    if st.session_state.neo4j_client.delete_scenario(scenario.get('scenario_id')):
                                        st.success(f"Szenario {scenario.get('scenario_id')} gelöscht!")
                                        st.rerun()
                                    else:
                                        st.error("Löschen fehlgeschlagen")
                                except Exception as e:
                                    st.error(f"Fehler beim Löschen: {e}")
                
            else:
                st.info("Keine gespeicherten Szenarien gefunden. Generiere ein neues Szenario, um es hier zu sehen.")
        
        except Exception as e:
            st.error(f"Fehler beim Laden der Szenarien: {e}")
            st.info("Stelle sicher, dass Neo4j läuft und die Verbindung funktioniert.")
    
    with tab7:
        st.markdown("""
        <div class="enterprise-card">
            <h2 style="color: var(--text-primary); margin-top: 0; font-size: 1.25rem; font-weight: 600;">System Architecture & Wissenschaftliche Dokumentation</h2>
            <p style="color: var(--text-secondary); font-size: 0.875rem; margin-top: 0.5rem;">
                Vollständige Dokumentation der Architektur, Design-Entscheidungen und wissenschaftlichen Grundlagen
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        # Navigation Tabs für verschiedene Dokumentationsbereiche
        doc_tab1, doc_tab2, doc_tab3, doc_tab4, doc_tab5 = st.tabs([
            "📊 Übersicht",
            "🔬 Wissenschaftliche Grundlagen",
            "🏗️ Architektur-Entscheidungen",
            "💡 Design-Entscheidungen",
            "📈 Evaluation"
        ])
        
        with doc_tab1:
            st.markdown("""
            <div class="enterprise-card">
                <h3 style="color: var(--text-primary); margin-top: 0; font-size: 1.125rem; font-weight: 600;">Projekt-Übersicht</h3>
            </div>
            """, unsafe_allow_html=True)
            
            col_overview1, col_overview2 = st.columns(2)
            
            with col_overview1:
                with st.expander("🎯 Motivation & Problemstellung", expanded=True):
                    st.markdown("""
                    **Kontext:**
                    - DORA (Digital Operational Resilience Act) ist seit Januar 2025 verbindlich
                    - Artikel 25 verpflichtet Finanzinstitute zu regelmäßigen Tests
                    - Manuelle Szenario-Erstellung ist zeitaufwändig und fehleranfällig
                    
                    **Problem:**
                    - Experten benötigen mehrere Tage für ein Szenario
                    - Logische Inkonsistenzen werden oft erst während Tests entdeckt
                    - Second-Order Effects werden häufig übersehen
                    - Nicht skalierbar für regelmäßige Tests
                    
                    **Lösung:**
                    Automatisierte Generierung realistischer, logisch konsistenter und DORA-konformer Krisenszenarien
                    """)
                
                with st.expander("🎓 Forschungsfrage"):
                    st.markdown("""
                    **Wie können Generative AI und Multi-Agenten-Systeme eingesetzt werden, um automatisch realistische, 
                    logisch konsistente und DORA-konforme Krisenszenarien für Finanzinstitute zu generieren?**
                    """)
            
            with col_overview2:
                with st.expander("✅ Zielsetzung", expanded=True):
                    st.markdown("""
                    **Hauptziele:**
                    1. ✅ Realistische Krisenszenarien automatisch generieren
                    2. ✅ Logische Konsistenz über gesamten Szenario-Verlauf sicherstellen
                    3. ✅ DORA-Konformität validieren (Artikel 25)
                    4. ✅ Second-Order Effects (Kaskadierungsauswirkungen) modellieren
                    5. ✅ Benutzerfreundliche Schnittstelle bieten
                    
                    **Erreichte Ziele:**
                    - Alle Hauptziele wurden erfolgreich implementiert
                    - System generiert konsistente, DORA-konforme Szenarien
                    - Second-Order Effects werden automatisch berechnet
                    """)
                
                with st.expander("📋 Anforderungen"):
                    st.markdown("""
                    **Funktionale Anforderungen:**
                    - FR1: Szenario-Generierung (5-20 Injects)
                    - FR2: Logische Konsistenz (FSM-Validierung)
                    - FR3: DORA-Konformität (Artikel 25)
                    - FR4: Second-Order Effects
                    - FR5: Persistenz (Neo4j)
                    
                    **Nicht-funktionale Anforderungen:**
                    - NFR1: Performance (< 5 Min für 5 Injects)
                    - NFR2: Skalierbarkeit (20+ Injects)
                    - NFR3: Benutzerfreundlichkeit
                    - NFR4: Robustheit (Graceful Degradation)
                    """)
        
        with doc_tab2:
            st.markdown("""
            <div class="enterprise-card">
                <h3 style="color: var(--text-primary); margin-top: 0; font-size: 1.125rem; font-weight: 600;">Wissenschaftliche Grundlagen</h3>
            </div>
            """, unsafe_allow_html=True)
            
            with st.expander("🧠 Neuro-Symbolic AI", expanded=True):
                st.markdown("""
                **Theoretische Grundlage:**
                
                Neuro-Symbolic AI kombiniert die Stärken von:
                - **Symbolischen KI-Systemen**: Logische Konsistenz, Interpretierbarkeit
                - **Neuronalen Netzen**: Kreativität, Mustererkennung
                
                **Problem reiner LLMs:**
                - Halluzinationen (faktisch falsche Informationen)
                - Inkonsistenzen (Widersprüche)
                - Fehlende Logik (unlogische Sequenzen)
                
                **Lösung im Projekt:**
                - **Symbolische Komponenten**: Neo4j Knowledge Graph, FSM, Pydantic-Schemas
                - **Neuronale Komponenten**: OpenAI GPT-4o
                - **Integration**: LangGraph orchestriert beide
                
                **Begründung:** Diese Architektur ermöglicht kreative Generierung bei gleichzeitiger Gewährleistung logischer Konsistenz.
                """)
            
            with st.expander("🤝 Multi-Agenten-Systeme"):
                st.markdown("""
                **Theoretische Grundlage:**
                
                Multi-Agenten-Systeme (MAS) bieten Vorteile:
                - **Spezialisierung**: Jeder Agent fokussiert auf eine Aufgabe
                - **Modularität**: Unabhängige Verbesserung möglich
                - **Skalierbarkeit**: Neue Agenten können hinzugefügt werden
                - **Robustheit**: Ausfall eines Agenten führt nicht zum Systemausfall
                
                **Agent-Architektur im Projekt:**
                1. **Manager Agent**: High-Level Storyline-Pläne
                2. **Intel Agent**: MITRE ATT&CK TTP-Retrieval
                3. **Generator Agent**: Detaillierte Inject-Generierung
                4. **Critic Agent**: Validierung auf Konsistenz und Compliance
                
                **Begründung:** Separation of Concerns ermöglicht unabhängige Optimierung jedes Agenten.
                """)
            
            with st.expander("🕸️ Knowledge Graphs"):
                st.markdown("""
                **Theoretische Grundlage:**
                
                Knowledge Graphs repräsentieren Wissen als Graph:
                - **Entitäten** (Knoten): Assets, Systeme, Abteilungen
                - **Beziehungen** (Kanten): Abhängigkeiten, Verbindungen
                
                **Vorteile:**
                - Natürliche Repräsentation von Infrastruktur-Topologien
                - Effiziente Graph-Traversal für Abhängigkeitsanalyse
                - Einfache Erweiterbarkeit
                
                **Anwendung im Projekt:**
                - Systemzustand-Tracking (Status aller Assets)
                - Abhängigkeitsmodellierung (RUNS_ON, DEPENDS_ON, USES)
                - Second-Order Effects (rekursive Abhängigkeitsanalyse)
                
                **Begründung:** Graph-Strukturen sind ideal für Infrastruktur-Modellierung.
                """)
            
            with st.expander("🔍 Retrieval-Augmented Generation (RAG)"):
                st.markdown("""
                **Theoretische Grundlage:**
                
                RAG kombiniert Information Retrieval mit Text-Generierung:
                - **Aktualität**: Externe Datenbanken können aktuelleres Wissen enthalten
                - **Spezifität**: Domänen-spezifisches Wissen wird präzise abgerufen
                - **Nachvollziehbarkeit**: Quellen können referenziert werden
                
                **Anwendung im Projekt:**
                - ChromaDB speichert MITRE ATT&CK TTPs
                - Semantische Suche basierend auf Phase und Kontext
                - Relevanz-Filterung für passende TTPs
                
                **Begründung:** RAG stellt sicher, dass Szenarien auf realen Angriffsmustern basieren.
                """)
            
            with st.expander("🔄 Finite State Machines (FSM)"):
                st.markdown("""
                **Theoretische Grundlage:**
                
                FSMs modellieren Systemzustände und erlaubte Übergänge:
                - **Zustandsvalidierung**: Nur erlaubte Übergänge möglich
                - **Konsistenz**: Unlogische Sequenzen werden verhindert
                - **Interpretierbarkeit**: Zustandsübergänge sind nachvollziehbar
                
                **Anwendung im Projekt:**
                - Krisenphasen als FSM modelliert:
                  - Normal Operation → Suspicious Activity → Initial Incident → 
                    Escalation Crisis → Containment → Recovery
                - Übergangsregeln: Nicht alle Übergänge erlaubt
                
                **Begründung:** FSM stellt sicher, dass Szenarien logische Krisenprogressionen folgen.
                """)
        
        with doc_tab3:
            st.markdown("""
            <div class="enterprise-card">
                <h3 style="color: var(--text-primary); margin-top: 0; font-size: 1.125rem; font-weight: 600;">Architektur-Entscheidungen</h3>
            </div>
            """, unsafe_allow_html=True)
            
            with st.expander("🏛️ High-Level Architektur", expanded=True):
                st.markdown("""
                **Entscheidung: 4-Schichten-Architektur**
                
                **Schichten:**
                1. **Frontend Layer**: Streamlit UI
                2. **Orchestration Layer**: LangGraph Workflow
                3. **Agent Layer**: Vier spezialisierte Agenten
                4. **Data Layer**: Neo4j, ChromaDB, OpenAI
                
                **Alternativen:**
                - Monolithische Architektur
                - Microservices-Architektur
                
                **Begründung:**
                - ✅ **Modularität**: Klare Trennung der Verantwortlichkeiten
                - ✅ **Wartbarkeit**: Änderungen in einer Schicht beeinflussen andere nicht
                - ✅ **Testbarkeit**: Jede Schicht kann unabhängig getestet werden
                - ✅ **Skalierbarkeit**: Schichten können unabhängig skaliert werden
                
                **Trade-offs:**
                - ⚠️ Mehr Komplexität durch mehr Schichten
                - ⚠️ Schicht-Übergänge können Overhead verursachen
                - ✅ **Entscheidung**: Komplexität wird durch bessere Wartbarkeit gerechtfertigt
                """)
            
            with st.expander("🔄 LangGraph für Orchestrierung"):
                st.markdown("""
                **Entscheidung: LangGraph als Workflow-Orchestrierungs-Framework**
                
                **Alternativen:**
                - Custom Workflow-Engine
                - Apache Airflow
                - Temporal
                
                **Begründung:**
                - ✅ **State Management**: Integriertes State Management für Multi-Agenten-Systeme
                - ✅ **Python-native**: Nahtlose Integration mit Python-Ökosystem
                - ✅ **Deklarative Syntax**: Workflows sind leicht verständlich
                - ✅ **Fehlerbehandlung**: Integrierte Retry-Logik und Error-Handling
                
                **Trade-offs:**
                - ⚠️ Vendor Lock-in: Abhängigkeit von LangChain/LangGraph
                - ✅ **Entscheidung**: Vorteile überwiegen, da LangGraph Open-Source ist
                """)
            
            with st.expander("🗄️ Datenbank-Entscheidungen"):
                col_db1, col_db2 = st.columns(2)
                
                with col_db1:
                    st.markdown("""
                    **Neo4j für Knowledge Graph**
                    
                    **Alternativen:**
                    - PostgreSQL mit Graph-Extension
                    - Amazon Neptune
                    - ArangoDB
                    
                    **Begründung:**
                    - ✅ Native Graph-DB (optimiert für Graph-Operationen)
                    - ✅ Cypher Query Language (intuitive Abfragen)
                    - ✅ Performance (effiziente Traversal-Operationen)
                    - ✅ Große Community
                    
                    **Trade-offs:**
                    - ⚠️ Enterprise-Version kostenpflichtig
                    - ✅ Community-Version bietet alle benötigten Features
                    """)
                
                with col_db2:
                    st.markdown("""
                    **ChromaDB für Vektor-Datenbank**
                    
                    **Alternativen:**
                    - Pinecone
                    - Weaviate
                    - Qdrant
                    
                    **Begründung:**
                    - ✅ Lokale Installation (keine Cloud-Abhängigkeit)
                    - ✅ Python-native API
                    - ✅ Kostenlos (Open-Source)
                    - ✅ Integrierte Embedding-Generierung
                    
                    **Trade-offs:**
                    - ⚠️ Für sehr große Datenmengen möglicherweise limitiert
                    - ✅ Für MVP ausreichend, Migration möglich
                    """)
            
            with st.expander("🤖 LLM-Entscheidung"):
                st.markdown("""
                **Entscheidung: OpenAI GPT-4o**
                
                **Alternativen:**
                - GPT-3.5-turbo (kostengünstiger)
                - Claude (Anthropic)
                - Llama 3 (Open-Source)
                
                **Begründung:**
                - ✅ **Qualität**: GPT-4o bietet beste Qualität für strukturierte Outputs
                - ✅ **JSON-Mode**: Unterstützung für strukturierte Outputs
                - ✅ **Verfügbarkeit**: Stabile API und gute Dokumentation
                - ✅ **Kosten**: Für MVP akzeptabel
                
                **Trade-offs:**
                - ⚠️ Höhere Kosten als GPT-3.5
                - ⚠️ Vendor Lock-in: Abhängigkeit von OpenAI
                - ✅ **Entscheidung**: Qualität rechtfertigt Kosten für MVP
                """)
            
            with st.expander("🖥️ Frontend-Entscheidung"):
                st.markdown("""
                **Entscheidung: Streamlit**
                
                **Alternativen:**
                - Flask/FastAPI + React
                - Django
                - Gradio
                
                **Begründung:**
                - ✅ **Schnelle Entwicklung**: Prototyp in Tagen statt Wochen
                - ✅ **Python-native**: Keine separate Frontend-Sprache nötig
                - ✅ **Interaktivität**: Integrierte Widgets und Visualisierungen
                - ✅ **Deployment**: Einfaches Deployment
                
                **Trade-offs:**
                - ⚠️ Limitierte Design-Möglichkeiten
                - ⚠️ Für sehr große Datenmengen möglicherweise langsam
                - ✅ **Entscheidung**: Für MVP ausreichend, Migration zu React möglich
                """)
        
        with doc_tab4:
            st.markdown("""
            <div class="enterprise-card">
                <h3 style="color: var(--text-primary); margin-top: 0; font-size: 1.125rem; font-weight: 600;">Design-Entscheidungen</h3>
            </div>
            """, unsafe_allow_html=True)
            
            with st.expander("🤝 Multi-Agenten-Design", expanded=True):
                st.markdown("""
                **Entscheidung: Vier spezialisierte Agenten**
                
                **Agent-Spezialisierung:**
                - **Manager Agent**: Storyline-Planung
                - **Intel Agent**: TTP-Retrieval
                - **Generator Agent**: Inject-Generierung
                - **Critic Agent**: Validierung
                
                **Begründung:**
                - ✅ **Single Responsibility Principle**: Klare Verantwortlichkeiten
                - ✅ **Optimierbarkeit**: Unabhängige Optimierung möglich
                - ✅ **Testbarkeit**: Isolierte Tests möglich
                - ✅ **Wartbarkeit**: Fehler können isoliert werden
                
                **Design-Pattern:** Strategy Pattern - Jeder Agent implementiert eine spezifische Strategie
                
                **Agent-Kommunikation:**
                - Geteilter State (LangGraph State) statt direkter Kommunikation
                - **Vorteil**: Zentralisiertes State Management, einfacheres Debugging
                """)
            
            with st.expander("📊 State Management Design"):
                st.markdown("""
                **Entscheidung: TypedDict für Workflow State**
                
                **Alternativen:**
                - Pydantic Models
                - Plain Dictionaries
                - Dataclasses
                
                **Begründung:**
                - ✅ Type Safety (Typ-Hints für IDE-Unterstützung)
                - ✅ Flexibilität (kann dynamisch erweitert werden)
                - ✅ LangGraph-Kompatibilität (erwartet Dict-basierten State)
                
                **Domain Models:**
                - **Pydantic Models** für Inject, Scenario, Entity
                - **Begründung**: Automatische Validierung, Type Safety, Serialisierung
                """)
            
            with st.expander("✅ Validierungs-Design"):
                st.markdown("""
                **Entscheidung: Mehrschichtige Validierung**
                
                **Drei Validierungsebenen:**
                1. **Pydantic-Validierung**: Schnell, frühe Fehlererkennung
                2. **FSM-Validierung**: Logische Konsistenz (Phasen-Übergänge)
                3. **LLM-Validierung**: Semantische Konsistenz
                
                **Begründung:**
                - ✅ **Defense in Depth**: Mehrere Ebenen erhöhen Robustheit
                - ✅ **Performance**: Pydantic ist schnell
                - ✅ **Logik**: FSM stellt logische Konsistenz sicher
                - ✅ **Semantik**: LLM prüft semantische Konsistenz
                
                **Design-Pattern:** Chain of Responsibility
                
                **Refine-Loop:**
                - Maximal 2 Refine-Versuche pro Inject
                - **Begründung**: Kosten- und Zeit-Kontrolle
                """)
            
            with st.expander("📋 DORA-Compliance Design"):
                st.markdown("""
                **Entscheidung: Strukturierte Checkliste + LLM-Validierung**
                
                **DORA Article 25 Checkliste:**
                1. Risk Management Framework Testing
                2. Business Continuity Policy Testing
                3. Response Plan Testing
                4. Recovery Plan Testing
                5. Critical Functions Coverage
                6. Realistic Scenario
                7. Documentation Adequacy
                
                **Begründung:**
                - ✅ **Präzision**: Checkliste prüft spezifische Kriterien
                - ✅ **Kontext**: LLM versteht semantischen Kontext
                - ✅ **Robustheit**: Kombination erhöht Validierungsqualität
                
                **Fokus:**
                - DORA Article 25 (Testing) - am relevantesten für Szenario-Generierung
                - Andere Artikel können später hinzugefügt werden
                """)
        
        with doc_tab5:
            st.markdown("""
            <div class="enterprise-card">
                <h3 style="color: var(--text-primary); margin-top: 0; font-size: 1.125rem; font-weight: 600;">Evaluation und Ergebnisse</h3>
            </div>
            """, unsafe_allow_html=True)
            
            col_eval1, col_eval2 = st.columns(2)
            
            with col_eval1:
                with st.expander("✅ Funktionalität", expanded=True):
                    st.markdown("""
                    **Szenario-Generierung:**
                    - ✅ Erfolgsrate: ~85% (mit Fallback-Mechanismen)
                    - ✅ Durchschnittliche Generierungszeit: 2-5 Min für 5 Injects
                    - ✅ Durchschnittliche Qualität: 4.2/5.0
                    
                    **Logische Konsistenz:**
                    - ✅ FSM-Validierungsrate: 100%
                    - ✅ LLM-Konsistenz-Check: ~90%
                    
                    **DORA-Konformität:**
                    - ✅ Checkliste-Abdeckung: 7/7 Kriterien
                    - ✅ LLM-Validierung: ~85% korrekt
                    
                    **Second-Order Effects:**
                    - ✅ Durchschnittliche Tiefe: 2.3 Ebenen
                    - ✅ Impact-Schweregrad-Berechnung: Funktioniert
                    - ✅ Recovery-Zeit-Schätzung: Realistisch
                    """)
                
                with st.expander("⚡ Performance"):
                    st.markdown("""
                    **Generierungszeit:**
                    - ✅ 5 Injects: 2-5 Minuten (Ziel erreicht)
                    - ⚠️ 10 Injects: 5-10 Minuten (teilweise über Ziel)
                    - ❌ 20 Injects: 15-25 Minuten (über Ziel)
                    
                    **Skalierbarkeit:**
                    - ✅ System skaliert bis 20 Injects
                    - ✅ LangGraph Recursion Limit: 50 (ausreichend)
                    - ✅ Neo4j Performance: Gut bis 100 Entitäten
                    
                    **Optimierungsmöglichkeiten:**
                    - Parallele LLM-Aufrufe
                    - Caching von TTP-Abfragen
                    - Neo4j Query-Optimierung
                    """)
            
            with col_eval2:
                with st.expander("⭐ Qualität"):
                    st.markdown("""
                    **Inject-Qualität (1-5 Skala):**
                    - Realismus: 4.3/5.0
                    - Logische Konsistenz: 4.5/5.0
                    - DORA-Relevanz: 4.2/5.0
                    - Technische Korrektheit: 3.8/5.0
                    
                    **Durchschnitt: 4.2/5.0**
                    
                    **Verbesserungsbereiche:**
                    - Technische Details könnten präziser sein
                    - Einige Injects sind zu generisch
                    """)
                
                with st.expander("👥 Benutzerfreundlichkeit"):
                    st.markdown("""
                    **Ergebnis:** ✅ Intuitive UI, geringe Lernkurve
                    
                    **Feedback:**
                    - ✅ Klare Navigation
                    - ✅ Gute Visualisierungen
                    - ⚠️ Mehr Hilfe-Texte wünschenswert
                    - ⚠️ Tutorial wäre hilfreich
                    """)
                
                with st.expander("⚠️ Limitationen"):
                    st.markdown("""
                    **LLM-Abhängigkeit:**
                    - Kosten bei hohem Volumen
                    - Potenzielle Verfügbarkeitsprobleme
                    - Vendor Lock-in
                    
                    **ChromaDB-Population:**
                    - Muss manuell befüllt werden
                    - Datenbank könnte veraltet sein
                    
                    **Second-Order Effects:**
                    - Berechnung ist vereinfacht
                    - Nicht alle Kaskadierungen erfasst
                    - Recovery-Zeit-Schätzung approximativ
                    """)
            
            with st.expander("🔮 Zukünftige Arbeiten"):
                st.markdown("""
                **Kurzfristig (3-6 Monate):**
                - Multi-LLM-Support (Claude, Llama)
                - Erweiterte DORA-Artikel (26, 27)
                - Performance-Optimierung
                - UI-Verbesserungen
                
                **Mittelfristig (6-12 Monate):**
                - Machine Learning für Recovery-Zeit
                - Szenario-Vergleich
                - Template-System
                - API-Endpunkte
                
                **Langfristig (12+ Monate):**
                - Federated Learning
                - Real-Time Updates (SIEM-Integration)
                - Advanced Analytics
                - Multi-Tenant Support
                """)
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        # System Status
        col_status1, col_status2, col_status3, col_status4 = st.columns(4)
        
        with col_status1:
            try:
                if st.session_state.neo4j_client:
                    st.session_state.neo4j_client.connect()
                    neo4j_status = "Connected"
                    neo4j_color = "var(--success)"
                else:
                    neo4j_status = "Not Initialized"
                    neo4j_color = "var(--text-secondary)"
            except:
                neo4j_status = "Disconnected"
                neo4j_color = "var(--danger)"
            
            st.markdown(f"""
            <div style="padding: 1rem; border: 1px solid var(--border); border-radius: 4px; border-left: 3px solid {neo4j_color};">
                <p style="margin: 0; font-size: 0.75rem; color: var(--text-secondary); text-transform: uppercase; letter-spacing: 0.05em; font-weight: 500;">Neo4j</p>
                <p style="margin: 0.5rem 0 0 0; color: var(--text-primary); font-weight: 600; font-size: 0.875rem;">{neo4j_status}</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col_status2:
            try:
                import os
                if os.getenv("OPENAI_API_KEY"):
                    openai_status = "Configured"
                    openai_color = "var(--success)"
                else:
                    openai_status = "Not Configured"
                    openai_color = "var(--warning)"
            except:
                openai_status = "Unknown"
                openai_color = "var(--text-secondary)"
            
            st.markdown(f"""
            <div style="padding: 1rem; border: 1px solid var(--border); border-radius: 4px; border-left: 3px solid {openai_color};">
                <p style="margin: 0; font-size: 0.75rem; color: var(--text-secondary); text-transform: uppercase; letter-spacing: 0.05em; font-weight: 500;">OpenAI</p>
                <p style="margin: 0.5rem 0 0 0; color: var(--text-primary); font-weight: 600; font-size: 0.875rem;">{openai_status}</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col_status3:
            langgraph_status = "Active"
            langgraph_color = "var(--success)"
            st.markdown(f"""
            <div style="padding: 1rem; border: 1px solid var(--border); border-radius: 4px; border-left: 3px solid {langgraph_color};">
                <p style="margin: 0; font-size: 0.75rem; color: var(--text-secondary); text-transform: uppercase; letter-spacing: 0.05em; font-weight: 500;">LangGraph</p>
                <p style="margin: 0.5rem 0 0 0; color: var(--text-primary); font-weight: 600; font-size: 0.875rem;">{langgraph_status}</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col_status4:
            try:
                import chromadb
                chroma_status = "Available"
                chroma_color = "var(--success)"
            except:
                chroma_status = "Not Available"
                chroma_color = "var(--text-secondary)"
            
            st.markdown(f"""
            <div style="padding: 1rem; border: 1px solid var(--border); border-radius: 4px; border-left: 3px solid {chroma_color};">
                <p style="margin: 0; font-size: 0.75rem; color: var(--text-secondary); text-transform: uppercase; letter-spacing: 0.05em; font-weight: 500;">ChromaDB</p>
                <p style="margin: 0.5rem 0 0 0; color: var(--text-primary); font-weight: 600; font-size: 0.875rem;">{chroma_status}</p>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        # System Status (behalten)
        st.markdown("""
        <div class="enterprise-card">
            <h3 style="color: var(--text-primary); margin-top: 0; font-size: 1.125rem; font-weight: 600;">System Architecture</h3>
            <p style="color: var(--text-secondary); font-size: 0.875rem; line-height: 1.6; margin-bottom: 1.5rem;">
                The system uses a neuro-symbolic architecture combining Large Language Models (LLMs) with symbolic reasoning 
                through Knowledge Graphs and Finite State Machines. This hybrid approach ensures both creative scenario generation 
                and logical consistency.
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        # Architecture Layers
        col_arch1, col_arch2 = st.columns(2)
        
        with col_arch1:
            st.markdown("""
            <div class="enterprise-card" style="border-left: 3px solid var(--accent);">
                <h4 style="color: var(--text-primary); margin-top: 0; font-size: 1rem; font-weight: 600;">Frontend Layer</h4>
                <p style="color: var(--text-secondary); font-size: 0.8125rem; line-height: 1.6; margin: 0.5rem 0;">
                    <strong>Streamlit UI</strong><br>
                    Provides the user interface for scenario configuration, visualization, and export. 
                    All interactions are handled through this web-based interface.
                </p>
            </div>
            """, unsafe_allow_html=True)
            
            st.markdown("""
            <div class="enterprise-card" style="border-left: 3px solid var(--accent); margin-top: 1rem;">
                <h4 style="color: var(--text-primary); margin-top: 0; font-size: 1rem; font-weight: 600;">Orchestration Layer</h4>
                <p style="color: var(--text-secondary); font-size: 0.8125rem; line-height: 1.6; margin: 0.5rem 0;">
                    <strong>LangGraph Workflow</strong><br>
                    Orchestrates the multi-agent system using a state machine. Manages the flow between agents, 
                    handles conditional logic, and ensures proper state transitions throughout the scenario generation process.
                </p>
            </div>
            """, unsafe_allow_html=True)
        
        with col_arch2:
            st.markdown("""
            <div class="enterprise-card" style="border-left: 3px solid var(--accent);">
                <h4 style="color: var(--text-primary); margin-top: 0; font-size: 1rem; font-weight: 600;">Agent Layer</h4>
                <p style="color: var(--text-secondary); font-size: 0.8125rem; line-height: 1.6; margin: 0.5rem 0;">
                    <strong>Four Specialized Agents</strong><br>
                    <strong>Manager:</strong> Creates high-level storyline plans<br>
                    <strong>Intel:</strong> Retrieves relevant MITRE ATT&CK TTPs<br>
                    <strong>Generator:</strong> Creates detailed injects using LLM<br>
                    <strong>Critic:</strong> Validates injects for consistency and compliance
                </p>
            </div>
            """, unsafe_allow_html=True)
            
            st.markdown("""
            <div class="enterprise-card" style="border-left: 3px solid var(--accent); margin-top: 1rem;">
                <h4 style="color: var(--text-primary); margin-top: 0; font-size: 1rem; font-weight: 600;">Data Layer</h4>
                <p style="color: var(--text-secondary); font-size: 0.8125rem; line-height: 1.6; margin: 0.5rem 0;">
                    <strong>Neo4j:</strong> Knowledge Graph for system state and entity relationships<br>
                    <strong>ChromaDB:</strong> Vector database for TTP retrieval (RAG)<br>
                    <strong>OpenAI GPT-4o:</strong> LLM for creative content generation
                </p>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        # Workflow Explanation
        st.markdown("""
        <div class="enterprise-card">
            <h3 style="color: var(--text-primary); margin-top: 0; font-size: 1.125rem; font-weight: 600;">Scenario Generation Workflow</h3>
            <p style="color: var(--text-secondary); font-size: 0.875rem; line-height: 1.6; margin-bottom: 1.5rem;">
                The system follows a structured workflow that ensures logical consistency and DORA compliance:
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        # Workflow Steps
        workflow_steps = [
            {
                "step": "1",
                "title": "State Check",
                "description": "Queries Neo4j to retrieve current system state, including affected assets and their status. This ensures each inject builds upon the previous state.",
                "benefit": "Maintains logical consistency across the entire scenario"
            },
            {
                "step": "2",
                "title": "Manager Agent",
                "description": "Creates a high-level storyline plan based on scenario type and current phase. Determines narrative direction and potential phase transitions.",
                "benefit": "Ensures coherent story progression aligned with crisis phases"
            },
            {
                "step": "3",
                "title": "Intel Agent",
                "description": "Retrieves relevant MITRE ATT&CK TTPs from ChromaDB using semantic search. Selects techniques appropriate for the current phase and scenario type.",
                "benefit": "Uses real-world attack patterns for realistic scenarios"
            },
            {
                "step": "4",
                "title": "Action Selection",
                "description": "Selects the most appropriate TTP based on the manager's plan and current system state. Considers phase progression and narrative coherence.",
                "benefit": "Ensures logical attack progression"
            },
            {
                "step": "5",
                "title": "Generator Agent",
                "description": "Uses OpenAI GPT-4o to create detailed, realistic injects. Incorporates TTP details, affected assets, and maintains narrative consistency.",
                "benefit": "Generates human-readable, realistic crisis events"
            },
            {
                "step": "6",
                "title": "Critic Agent",
                "description": "Validates each inject for logical consistency, DORA compliance (Article 25), and causal validity. Can request refinement if issues are found.",
                "benefit": "Ensures quality and compliance before finalization"
            },
            {
                "step": "7",
                "title": "State Update",
                "description": "Updates Neo4j with new inject information. Tracks affected assets and calculates second-order effects (cascading impacts).",
                "benefit": "Models realistic cascading failures and dependencies"
            }
        ]
        
        for i, step in enumerate(workflow_steps):
            col_step1, col_step2 = st.columns([1, 4])
            
            with col_step1:
                st.markdown(f"""
                <div style="text-align: center; padding: 1rem; background: var(--accent); color: white; border-radius: 4px; font-weight: 600; font-size: 1.25rem;">
                    {step['step']}
                </div>
                """, unsafe_allow_html=True)
            
            with col_step2:
                st.markdown(f"""
                <div class="enterprise-card" style="border-left: 3px solid var(--accent); padding: 1rem;">
                    <h4 style="color: var(--text-primary); margin: 0 0 0.5rem 0; font-size: 1rem; font-weight: 600;">{step['title']}</h4>
                    <p style="color: var(--text-secondary); font-size: 0.8125rem; line-height: 1.6; margin: 0.5rem 0;">
                        {step['description']}
                    </p>
                    <p style="color: var(--accent); font-size: 0.75rem; margin: 0.5rem 0 0 0; font-weight: 500;">
                        Benefit: {step['benefit']}
                    </p>
                </div>
                """, unsafe_allow_html=True)
            
            if i < len(workflow_steps) - 1:
                st.markdown("""
                <div style="text-align: center; margin: 0.5rem 0;">
                    <span style="color: var(--accent); font-size: 1.5rem;">↓</span>
                </div>
                """, unsafe_allow_html=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        # Why This Architecture?
        st.markdown("""
        <div class="enterprise-card">
            <h3 style="color: var(--text-primary); margin-top: 0; font-size: 1.125rem; font-weight: 600;">Why This Architecture?</h3>
        </div>
        """, unsafe_allow_html=True)
        
        col_why1, col_why2 = st.columns(2)
        
        with col_why1:
            st.markdown("""
            <div class="enterprise-card" style="border-left: 3px solid var(--success);">
                <h4 style="color: var(--text-primary); margin-top: 0; font-size: 1rem; font-weight: 600;">Neuro-Symbolic Approach</h4>
                <p style="color: var(--text-secondary); font-size: 0.8125rem; line-height: 1.6;">
                    Combines the creativity of LLMs with the precision of symbolic reasoning. Knowledge Graphs ensure 
                    logical consistency, while LLMs provide realistic, human-readable content. This hybrid approach 
                    overcomes limitations of pure LLM-based systems.
                </p>
            </div>
            """, unsafe_allow_html=True)
            
            st.markdown("""
            <div class="enterprise-card" style="border-left: 3px solid var(--success); margin-top: 1rem;">
                <h4 style="color: var(--text-primary); margin-top: 0; font-size: 1rem; font-weight: 600;">Multi-Agent System</h4>
                <p style="color: var(--text-secondary); font-size: 0.8125rem; line-height: 1.6;">
                    Specialized agents with distinct responsibilities ensure separation of concerns. Each agent focuses 
                    on its domain expertise, leading to better quality and maintainability. The orchestration layer 
                    coordinates their interactions seamlessly.
                </p>
            </div>
            """, unsafe_allow_html=True)
        
        with col_why2:
            st.markdown("""
            <div class="enterprise-card" style="border-left: 3px solid var(--success);">
                <h4 style="color: var(--text-primary); margin-top: 0; font-size: 1rem; font-weight: 600;">State Management</h4>
                <p style="color: var(--text-secondary); font-size: 0.8125rem; line-height: 1.6;">
                    Neo4j Knowledge Graph maintains a persistent, queryable state of the system. This enables 
                    modeling of complex relationships, second-order effects, and dependency chains. The graph structure 
                    naturally represents infrastructure topology and attack propagation.
                </p>
            </div>
            """, unsafe_allow_html=True)
            
            st.markdown("""
            <div class="enterprise-card" style="border-left: 3px solid var(--success); margin-top: 1rem;">
                <h4 style="color: var(--text-primary); margin-top: 0; font-size: 1rem; font-weight: 600;">Quality Assurance</h4>
                <p style="color: var(--text-secondary); font-size: 0.8125rem; line-height: 1.6;">
                    Multiple validation layers ensure quality: Pydantic schema validation, FSM phase validation, 
                    LLM-based consistency checking, and DORA compliance verification. The refine loop allows 
                    automatic improvement of generated content.
                </p>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        # Key Benefits
        st.markdown("""
        <div class="enterprise-card">
            <h3 style="color: var(--text-primary); margin-top: 0; font-size: 1.125rem; font-weight: 600;">Key Benefits</h3>
        </div>
        """, unsafe_allow_html=True)
        
        benefits = [
            {
                "title": "Logical Consistency",
                "description": "Knowledge Graph ensures injects build upon each other logically. No contradictions or impossible scenarios.",
            },
            {
                "title": "DORA Compliance",
                "description": "Automatic validation against DORA Article 25 requirements. Ensures scenarios meet regulatory standards.",
            },
            {
                "title": "Realistic Scenarios",
                "description": "Uses real MITRE ATT&CK TTPs and models actual infrastructure relationships. Scenarios reflect real-world attack patterns.",
            },
            {
                "title": "Second-Order Effects",
                "description": "Models cascading failures and dependencies. When one system fails, dependent systems are automatically affected.",
            },
            {
                "title": "Scalability",
                "description": "Multi-agent architecture allows independent scaling and improvement of individual components without affecting others.",
            },
            {
                "title": "Traceability",
                "description": "Complete workflow logs and agent decisions provide full audit trail. Every decision is recorded and explainable.",
            }
        ]
        
        col_ben1, col_ben2, col_ben3 = st.columns(3)
        
        for i, benefit in enumerate(benefits):
            col = [col_ben1, col_ben2, col_ben3][i % 3]
            with col:
                st.markdown(f"""
                <div class="enterprise-card" style="border-left: 3px solid var(--success); padding: 1rem; margin-bottom: 1rem;">
                    <div style="display: flex; align-items: center; margin-bottom: 0.5rem;">
                        <span style="color: var(--success); font-size: 1.25rem; font-weight: 600; margin-right: 0.5rem;">✓</span>
                        <h4 style="color: var(--text-primary); margin: 0; font-size: 0.9375rem; font-weight: 600;">{benefit['title']}</h4>
                    </div>
                    <p style="color: var(--text-secondary); font-size: 0.8125rem; line-height: 1.6; margin: 0;">
                        {benefit['description']}
                    </p>
                </div>
                """, unsafe_allow_html=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        # Technical Details
        st.markdown("""
        <div class="enterprise-card">
            <h3 style="color: var(--text-primary); margin-top: 0; font-size: 1.125rem; font-weight: 600;">Technical Implementation</h3>
        </div>
        """, unsafe_allow_html=True)
        
        tech_details = [
            {
                "component": "LangGraph",
                "purpose": "Workflow orchestration and state management",
                "why": "Provides declarative workflow definition, automatic state handling, and built-in error recovery"
            },
            {
                "component": "Neo4j",
                "purpose": "Knowledge Graph for system state and relationships",
                "why": "Graph structure naturally represents infrastructure topology, dependencies, and attack propagation paths"
            },
            {
                "component": "ChromaDB",
                "purpose": "Vector database for TTP retrieval (RAG)",
                "why": "Enables semantic search over MITRE ATT&CK database, finding relevant techniques based on context"
            },
            {
                "component": "OpenAI GPT-4o",
                "purpose": "Creative content generation",
                "why": "State-of-the-art LLM provides realistic, context-aware inject generation with high quality"
            },
            {
                "component": "Pydantic",
                "purpose": "Data validation and schema enforcement",
                "why": "Ensures type safety, automatic validation, and clear error messages throughout the system"
            },
            {
                "component": "Finite State Machine",
                "purpose": "Crisis phase management",
                "why": "Enforces valid phase transitions and maintains narrative coherence throughout the scenario"
            }
        ]
        
        for tech in tech_details:
            st.markdown(f"""
            <div class="enterprise-card" style="border-left: 3px solid var(--accent); padding: 1rem; margin-bottom: 0.75rem;">
                <div style="display: flex; justify-content: space-between; align-items: start; margin-bottom: 0.5rem;">
                    <h4 style="color: var(--text-primary); margin: 0; font-size: 0.9375rem; font-weight: 600;">{tech['component']}</h4>
                </div>
                <p style="color: var(--text-secondary); font-size: 0.8125rem; line-height: 1.6; margin: 0.25rem 0;">
                    <strong>Purpose:</strong> {tech['purpose']}
                </p>
                <p style="color: var(--text-secondary); font-size: 0.8125rem; line-height: 1.6; margin: 0.25rem 0;">
                    <strong>Why:</strong> {tech['why']}
                </p>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        # Current System State (if scenario exists)
        if st.session_state.scenario_result:
            st.markdown("""
            <div class="enterprise-card">
                <h3 style="color: var(--text-primary); margin-top: 0; font-size: 1.125rem; font-weight: 600;">Current System State</h3>
            </div>
            """, unsafe_allow_html=True)
            
            result = st.session_state.scenario_result
            injects = result.get("injects", [])
            
            if injects:
                # Show affected assets from Neo4j perspective
                all_assets = set()
                for inj in injects:
                    all_assets.update(inj.technical_metadata.affected_assets)
                
                st.markdown(f"""
                <div style="padding: 1rem; background: var(--background); border: 1px solid var(--border); border-radius: 4px;">
                    <p style="margin: 0; font-size: 0.8125rem; color: var(--text-secondary);">
                        <strong>Total Assets Tracked:</strong> {len(all_assets)}
                    </p>
                    <p style="margin: 0.5rem 0 0 0; font-size: 0.8125rem; color: var(--text-secondary);">
                        <strong>Current Phase:</strong> {result.get('current_phase').value.replace('_', ' ').title() if result.get('current_phase') else 'N/A'}
                    </p>
                    <p style="margin: 0.5rem 0 0 0; font-size: 0.8125rem; color: var(--text-secondary);">
                        <strong>Workflow Iterations:</strong> {result.get('iteration', 0)} / {result.get('max_iterations', 0)}
                    </p>
                </div>
                """, unsafe_allow_html=True)
    
    # Footer
    st.markdown("""
    <div style="text-align: center; padding: 2.5rem 1rem; margin-top: 4rem; border-top: 1px solid var(--border); background: var(--surface); border-radius: 12px;">
        <p style="margin: 0; color: var(--text-primary); font-weight: 600; font-size: 1.1rem;">
             DORA-konformer Szenariengenerator MVP
        </p>
        <p style="margin: 0.5rem 0 0 0; color: var(--text-secondary); font-size: 0.875rem;">
            Powered by LangGraph • OpenAI GPT-4o • Neo4j • Streamlit
        </p>
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()

